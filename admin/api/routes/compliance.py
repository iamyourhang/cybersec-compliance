"""
admin/api/routes/compliance.py
合规知识库 CRUD 接口
"""

from __future__ import annotations

import json
import logging
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, UploadFile, File, Form
from pydantic import BaseModel

from admin.api.auth import get_current_user, is_admin_username, require_admin_user
from admin.api.routes.documents import _parse_and_index_document
from collector.audit.authenticity import assess_record_authenticity
from collector.document.source_ingest import OfficialSourceIngestService
from collector.query.service import get_compliance_query_service
from collector.review.service import get_authenticity_review_service
from collector.official_sources.repository import OfficialSourceRepository
from collector.translation.repository import list_translations_for_entities
from collector.translation.service import attach_translation_fields
from database.connection import get_cursor, get_connection
from database.repository import (
    CanonicalRequirementRepository,
    ChangeLogRepository,
    ComplianceIndexRepository,
    ComplianceRepository,
    ReviewCaseRepository,
    SourceArtifactRepository,
)
from urllib.parse import urlparse

logger = logging.getLogger(__name__)
router = APIRouter()


def _translation_map(entity_type: str, entity_ids: List[str]) -> Dict[tuple[str, str], str]:
    try:
        return list_translations_for_entities(entity_type, entity_ids)
    except Exception as exc:
        logger.warning("translation lookup skipped [%s]: %s", entity_type, exc)
        return {}


# ---- Schema ----

class ComplianceCreate(BaseModel):
    name: str
    entry_type: str
    mandatory: str = "mandatory"
    status: str = "active"
    country_code: str
    issuing_body: Optional[str] = None
    technical_standards: Optional[List[str]] = []
    regulation_basis: Optional[List[str]] = []
    effective_date: Optional[str] = None
    transition_end_date: Optional[str] = None
    validity_period: Optional[str] = None
    published_date: Optional[str] = None
    applicable_products: Optional[List[str]] = []
    scope_description: Optional[str] = None
    requirements: Optional[Dict[str, Any]] = None
    testing_bodies: Optional[List[str]] = []
    assessment_procedure: Optional[str] = None
    official_url: Optional[str] = None
    remarks: Optional[str] = None


class ComplianceUpdate(BaseModel):
    name: Optional[str] = None
    entry_type: Optional[str] = None
    mandatory: Optional[str] = None
    status: Optional[str] = None
    issuing_body: Optional[str] = None
    technical_standards: Optional[List[str]] = None
    regulation_basis: Optional[List[str]] = None
    effective_date: Optional[str] = None
    transition_end_date: Optional[str] = None
    validity_period: Optional[str] = None
    applicable_products: Optional[List[str]] = None
    scope_description: Optional[str] = None
    requirements: Optional[Dict[str, Any]] = None
    official_url: Optional[str] = None
    remarks: Optional[str] = None
    verified: Optional[bool] = None


class ManualSourceRequest(BaseModel):
    official_url: str
    artifact_url: Optional[str] = None
    evidence_note: str
    auto_parse: bool = True


class ManualReviewRequest(BaseModel):
    authenticity_status: str
    risk_score: int
    reasons: List[str]
    evidence_note: str
    source_download_status: Optional[str] = None
    source_download_error: Optional[str] = None


def _normalize_domain(domain: str) -> str:
    return (domain or "").strip().lower().removeprefix("www.")


def _is_allowed_official_url(url: str, allowed_domains: List[str]) -> bool:
    parsed = urlparse(url)
    domain = _normalize_domain(parsed.netloc)
    if parsed.scheme not in {"http", "https"} or not domain:
        return False
    for allowed in allowed_domains:
        normalized = _normalize_domain(allowed)
        if domain == normalized or domain.endswith(f".{normalized}"):
            return True
    return False


def _get_record_allowed_domains(record: Dict[str, Any]) -> List[str]:
    source_name = (record.get("data_source") or "").split("official_source:", 1)[-1] if "official_source:" in (record.get("data_source") or "") else None
    sources = OfficialSourceRepository().list_all(enabled_only=False)
    matched: List[str] = []
    for source in sources:
        if source.get("country_code") != record.get("country_code"):
            continue
        if source_name and source.get("name") != source_name:
            continue
        matched.extend(source.get("allowed_domains") or [])
    if matched:
        return matched

    fallback = [
        domain
        for source in sources
        if source.get("country_code") == record.get("country_code")
        for domain in (source.get("allowed_domains") or [])
    ]
    official_url = (record.get("official_url") or "").strip()
    if official_url:
        domain = _normalize_domain(urlparse(official_url).netloc)
        if domain:
            fallback.append(domain)
    return sorted({_normalize_domain(domain) for domain in fallback if domain})


def _serialize_record_dates(row: Dict[str, Any], fields: Optional[List[str]] = None) -> Dict[str, Any]:
    payload = dict(row)
    date_fields = fields or [
        "effective_date",
        "transition_end_date",
        "published_date",
        "created_at",
        "updated_at",
        "last_checked",
        "checked_at",
        "downloaded_at",
    ]
    for field in date_fields:
        if payload.get(field):
            payload[field] = str(payload[field])
    return payload


def _pending_review_bucket_clause(review_bucket: Optional[str]) -> Optional[str]:
    if not review_bucket:
        return None
    if review_bucket == "official_url_missing":
        return " AND ci.official_url IS NULL"
    if review_bucket == "domain_mismatch":
        return " AND COALESCE(rc.reasons, '[]'::jsonb) ? 'official_domain_mismatch'"
    if review_bucket == "404_like":
        return """
            AND (
                COALESCE(rc.reasons, '[]'::jsonb) ? 'official_url_not_found'
                OR COALESCE(rc.reasons, '[]'::jsonb) ? 'official_url_redirected_to_home'
                OR (
                    COALESCE(rc.source_download_status, 'pending') = 'failed'
                    AND (
                        COALESCE(rc.source_download_error, '') ILIKE '%%404%%'
                        OR COALESCE(rc.source_download_error, '') ILIKE '%%not found%%'
                        OR COALESCE(rc.source_download_error, '') ILIKE '%%page not found%%'
                        OR COALESCE(rc.source_download_error, '') ILIKE '%%portal%%'
                        OR COALESCE(rc.source_download_error, '') ILIKE '%%homepage%%'
                    )
                )
            )
        """
    if review_bucket == "local_adoption_unverified":
        return " AND COALESCE(rc.reasons, '[]'::jsonb) ? 'local_adoption_unverified'"
    raise HTTPException(status_code=400, detail="不支持的 review_bucket")


# ---- 接口 ----


@router.get("/")
async def list_compliance(
    country_code: Optional[str] = None,
    entry_type: Optional[str] = None,
    mandatory: Optional[str] = None,
    status: Optional[str] = "active",
    product_code: Optional[str] = None,
    keyword: Optional[str] = None,
    verified: Optional[bool] = None,
    authenticity_status: Optional[str] = None,
    sort_by: Optional[str] = Query("updated_at", regex="^(name|entry_type|mandatory|country_code|effective_date|status|confidence_score|updated_at)$"),
    sort_order: Optional[str] = Query("desc", regex="^(asc|desc)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: str = Depends(get_current_user),
):
    """分页查询合规条目，支持多维度筛选和排序"""
    is_admin = is_admin_username(current_user)
    include_suspicious = False
    effective_authenticity_status = authenticity_status
    if not is_admin:
        effective_authenticity_status = "verified"
        include_suspicious = False
        verified = True
        authenticity_status = "verified"
    if not authenticity_status:
        effective_authenticity_status = "verified"
    if verified is True and not authenticity_status:
        effective_authenticity_status = "verified"
        include_suspicious = False
    elif verified is False and not authenticity_status:
        effective_authenticity_status = None
        include_suspicious = True

    data = get_compliance_query_service().list_compliance(
        country_code=country_code,
        entry_type=entry_type,
        mandatory=mandatory,
        status=status,
        product_code=product_code,
        keyword=keyword,
        authenticity_status=effective_authenticity_status,
        include_suspicious=include_suspicious,
        page=page,
        page_size=page_size,
        sort_by=sort_by,
        sort_order=sort_order,
    )
    items = []
    for row in data["items"]:
        item = dict(row)
        if item.get("compliance_id"):
            item["index_id"] = str(item.get("id")) if item.get("id") else None
            item["id"] = str(item["compliance_id"])
        for date_field in ["effective_date", "published_date", "updated_at"]:
            if item.get(date_field):
                item[date_field] = str(item[date_field])
        items.append(item)
    translations = _translation_map(
        "compliance_index",
        [str(item.get("id")) for item in items],
    )
    items = [
        attach_translation_fields(item, translations, entity_id_field="id")
        for item in items
    ]
    return {"total": data["total"], "page": page, "page_size": page_size, "items": items}


@router.get("/audit/suspicious")
async def list_suspicious_compliance(
    country_code: Optional[str] = None,
    entry_type: Optional[str] = None,
    limit: int = Query(50, ge=1, le=200),
    current_user: str = Depends(require_admin_user),
):
    sql = """
        SELECT ci.compliance_id AS id,
               ci.name,
               ci.country_code,
               ci.entry_type,
               ci.issuing_body,
               ci.official_url,
               (ci.authenticity_status = 'verified') AS verified,
               GREATEST(0, 100 - COALESCE(ci.authenticity_risk_score, 0)) AS confidence_score,
               COALESCE(sr.discovery_method, 'compliance_index') AS data_source,
               ci.status
        FROM compliance_index ci
        LEFT JOIN source_records sr ON sr.id = ci.source_record_id
        WHERE ci.status = 'active'
    """
    params: list = []
    if country_code:
        sql += " AND ci.country_code = %s"
        params.append(country_code)
    if entry_type:
        sql += " AND ci.entry_type = %s"
        params.append(entry_type)
    sql += " ORDER BY ci.updated_at DESC LIMIT %s"
    params.append(limit)

    with get_cursor() as cur:
        cur.execute(sql, params)
        rows = [dict(row) for row in cur.fetchall()]

    audited = []
    for row in rows:
        verdict = assess_record_authenticity(row)
        if verdict["risk_level"] in {"medium", "high", "critical"}:
            audited.append({**row, **verdict})

    audited.sort(key=lambda item: item["risk_score"], reverse=True)
    return {"items": audited, "total": len(audited)}


@router.post("/audit/quarantine-critical")
async def quarantine_critical_compliance(
    country_code: Optional[str] = None,
    entry_type: Optional[str] = None,
    limit: int = Query(200, ge=1, le=2000),
    current_user: str = Depends(require_admin_user),
):
    sql = """
        SELECT ci.compliance_id AS id,
               ci.name,
               ci.country_code,
               ci.entry_type,
               ci.issuing_body,
               ci.official_url,
               (ci.authenticity_status = 'verified') AS verified,
               GREATEST(0, 100 - COALESCE(ci.authenticity_risk_score, 0)) AS confidence_score,
               COALESCE(sr.discovery_method, 'compliance_index') AS data_source,
               ci.status
        FROM compliance_index ci
        LEFT JOIN source_records sr ON sr.id = ci.source_record_id
        WHERE ci.status = 'active'
          AND COALESCE(ci.authenticity_status, 'candidate') <> 'verified'
    """
    params: list = []
    if country_code:
        sql += " AND ci.country_code = %s"
        params.append(country_code)
    if entry_type:
        sql += " AND ci.entry_type = %s"
        params.append(entry_type)
    sql += " ORDER BY ci.updated_at DESC LIMIT %s"
    params.append(limit)

    with get_cursor() as cur:
        cur.execute(sql, params)
        rows = [dict(row) for row in cur.fetchall()]

    quarantined = []
    reviewed = 0
    for row in rows:
        verdict = assess_record_authenticity(row)
        status = "quarantined" if verdict["recommended_action"] == "quarantine" else "suspicious"
        evidence = ", ".join(verdict["reasons"])
        ComplianceRepository.set_authenticity_review(
            str(row["id"]),
            authenticity_status=status,
            risk_score=verdict["risk_score"],
            reasons=verdict["reasons"],
            checked_by=current_user,
            evidence=evidence,
        )
        reviewed += 1
        if status == "quarantined":
            quarantined.append(
                {
                    "id": str(row["id"]),
                    "name": row["name"],
                    "country_code": row["country_code"],
                    "risk_score": verdict["risk_score"],
                }
            )

    return {
        "reviewed": reviewed,
        "quarantined": quarantined,
        "quarantined_count": len(quarantined),
    }


@router.post("/{record_id}/source/download")
async def download_official_source(
    record_id: str,
    background_tasks: BackgroundTasks,
    auto_parse: bool = True,
    current_user: str = Depends(require_admin_user),
):
    record = ComplianceRepository.get_by_id(record_id)
    if not record:
        raise HTTPException(status_code=404, detail="条目不存在")

    try:
        result = OfficialSourceIngestService().ingest_record(dict(record), requested_by=current_user)
    except ValueError as exc:
        ComplianceRepository.mark_source_download_failed(record_id, str(exc))
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.error("官方原文抓取失败 [%s]: %s", record_id, exc, exc_info=True)
        ComplianceRepository.mark_source_download_failed(record_id, str(exc))
        raise HTTPException(status_code=502, detail="官方原文下载失败") from exc

    if auto_parse:
        background_tasks.add_task(_parse_and_index_document, result["doc_id"], True)

    return {
        **result,
        "auto_parse": auto_parse,
    }


@router.post("/{record_id}/manual-source")
async def manual_source(
    record_id: str,
    req: ManualSourceRequest,
    background_tasks: BackgroundTasks,
    current_user: str = Depends(require_admin_user),
):
    record = ComplianceRepository.get_by_id(record_id)
    if not record:
        raise HTTPException(status_code=404, detail="条目不存在")

    allowed_domains = _get_record_allowed_domains(dict(record))
    if not allowed_domains:
        raise HTTPException(status_code=400, detail="未找到该条目对应的官方域名白名单")
    if not _is_allowed_official_url(req.official_url, allowed_domains):
        raise HTTPException(status_code=400, detail="official_url 不在官方域名白名单内")
    if req.artifact_url and not _is_allowed_official_url(req.artifact_url, allowed_domains):
        raise HTTPException(status_code=400, detail="artifact_url 不在官方域名白名单内")

    try:
        result = OfficialSourceIngestService().ingest_manual_source(
            dict(record),
            official_url=req.official_url,
            artifact_url=req.artifact_url,
            requested_by=current_user,
        )
    except ValueError as exc:
        ComplianceRepository.mark_source_download_failed(record_id, str(exc))
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.error("人工补源失败 [%s]: %s", record_id, exc, exc_info=True)
        ComplianceRepository.mark_source_download_failed(record_id, str(exc))
        raise HTTPException(status_code=502, detail="人工补源失败") from exc

    get_authenticity_review_service().register_manual_source(
        dict(record),
        ingest_result=result,
        official_url=req.official_url,
        evidence_note=req.evidence_note,
        checked_by=current_user,
    )

    should_parse = bool(req.auto_parse and result.get("file_type") == "pdf")
    if should_parse:
        background_tasks.add_task(_parse_and_index_document, result["doc_id"], False)

    return {
        **result,
        "authenticity_status": "verified",
        "auto_parse": should_parse,
    }


@router.post("/{record_id}/manual-source-upload")
async def manual_source_upload(
    record_id: str,
    background_tasks: BackgroundTasks,
    official_url: str = Form(...),
    artifact_url: Optional[str] = Form(None),
    evidence_note: str = Form(...),
    auto_parse: bool = Form(True),
    file: UploadFile = File(...),
    current_user: str = Depends(require_admin_user),
):
    record = ComplianceRepository.get_by_id(record_id)
    if not record:
        raise HTTPException(status_code=404, detail="条目不存在")

    allowed_domains = _get_record_allowed_domains(dict(record))
    if not allowed_domains:
        raise HTTPException(status_code=400, detail="未找到该条目对应的官方域名白名单")
    if not _is_allowed_official_url(official_url, allowed_domains):
        raise HTTPException(status_code=400, detail="official_url 不在官方域名白名单内")
    if artifact_url and not _is_allowed_official_url(artifact_url, allowed_domains):
        raise HTTPException(status_code=400, detail="artifact_url 不在官方域名白名单内")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")
    content_type = file.content_type or ""
    lower_name = (file.filename or "").lower()
    if "pdf" not in content_type.lower() and "html" not in content_type.lower() and not (
        lower_name.endswith(".pdf") or lower_name.endswith(".html") or lower_name.endswith(".htm")
    ):
        raise HTTPException(status_code=400, detail="仅支持上传官方 PDF 或 HTML 正文页")

    try:
        result = OfficialSourceIngestService().ingest_uploaded_source(
            dict(record),
            official_url=official_url,
            file_name=file.filename or "official_source.html",
            content=content,
            content_type=content_type,
            artifact_url=artifact_url,
            requested_by=current_user,
        )
    except Exception as exc:
        logger.error("本地补源上传失败 [%s]: %s", record_id, exc, exc_info=True)
        ComplianceRepository.mark_source_download_failed(record_id, str(exc))
        raise HTTPException(status_code=502, detail="本地补源上传失败") from exc

    get_authenticity_review_service().register_manual_source(
        dict(record),
        ingest_result=result,
        official_url=official_url,
        evidence_note=evidence_note,
        checked_by=current_user,
    )

    should_parse = bool(auto_parse)
    if should_parse:
        background_tasks.add_task(_parse_and_index_document, result["doc_id"], False)

    return {
        **result,
        "authenticity_status": "verified",
        "auto_parse": should_parse,
    }


@router.post("/{record_id}/review")
async def manual_review(
    record_id: str,
    req: ManualReviewRequest,
    current_user: str = Depends(require_admin_user),
):
    record = ComplianceRepository.get_by_id(record_id)
    if not record:
        raise HTTPException(status_code=404, detail="条目不存在")

    if req.authenticity_status not in {"suspicious", "quarantined"}:
        raise HTTPException(
            status_code=400,
            detail="人工审核接口只允许写入 suspicious/quarantined；verified 请使用 /manual-source",
        )
    if not req.reasons:
        raise HTTPException(status_code=400, detail="reasons 不能为空")
    if not req.evidence_note.strip():
        raise HTTPException(status_code=400, detail="evidence_note 不能为空")
    if not 0 <= req.risk_score <= 100:
        raise HTTPException(status_code=400, detail="risk_score 必须在 0-100 之间")
    if req.source_download_status and req.source_download_status not in {"pending", "downloaded", "failed"}:
        raise HTTPException(status_code=400, detail="source_download_status 不合法")

    result = get_authenticity_review_service().apply_decision(
        case_id=ReviewCaseRepository.ensure_for_record(dict(record)),
        decision=req.model_dump(),
        checked_by=current_user,
    )

    return {
        "id": record_id,
        "authenticity_status": result.get("current_status", req.authenticity_status),
        "risk_score": req.risk_score,
        "source_download_status": req.source_download_status,
    }




# ============================================================
# 导出接口（独立追加）
# ============================================================


@router.get("/meta/all")
async def get_meta_all(current_user: str = Depends(get_current_user)):
    """一次性返回国家和产品元数据"""
    with get_cursor() as cur:
        cur.execute("SELECT code, name_zh, priority FROM countries WHERE enabled=TRUE ORDER BY priority, name_zh")
        countries = [dict(r) for r in cur.fetchall()]
        cur.execute("SELECT code, name_zh FROM products WHERE enabled=TRUE ORDER BY name_zh")
        products = [dict(r) for r in cur.fetchall()]
    return {"countries": countries, "products": products}


@router.get("/meta/countries")
async def get_countries(current_user: str = Depends(get_current_user)):
    with get_cursor() as cur:
        cur.execute("SELECT code, name_zh, priority FROM countries WHERE enabled=TRUE ORDER BY priority, name_zh")
        return [dict(r) for r in cur.fetchall()]


@router.get("/meta/products")
async def get_products(current_user: str = Depends(get_current_user)):
    with get_cursor() as cur:
        cur.execute("SELECT code, name_zh FROM products WHERE enabled=TRUE ORDER BY name_zh")
        return [dict(r) for r in cur.fetchall()]


@router.get("/export/excel")
async def export_excel(
    country_code: Optional[str] = None,
    entry_type: Optional[str] = None,
    mandatory: Optional[str] = None,
    status: Optional[str] = "active",
    product_code: Optional[str] = None,
    keyword: Optional[str] = None,
    verified: Optional[bool] = None,
    current_user: str = Depends(get_current_user),
):
    """按筛选条件导出已核实合规知识库 Excel。"""
    import io
    from datetime import date
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    # 查询数据
    sql = """
        SELECT
            ci.compliance_id AS id,
            ci.name,
            ci.entry_type,
            ci.mandatory,
            ci.country_code,
            ci.status,
            ci.issuing_body,
            ci.official_url,
            ci.applicable_products,
            ci.effective_date,
            ci.published_date,
            GREATEST(0, 100 - COALESCE(ci.authenticity_risk_score, 0)) AS confidence_score,
            TRUE AS verified,
            ci.authenticity_status,
            ci.authenticity_risk_score,
            ci.summary,
            ci.document_id,
            COALESCE(sa.artifact_url, sa.official_url) AS source_artifact_url,
            sa.artifact_sha256,
            ci.updated_at,
            c.name_zh AS country_name,
            c.priority
        FROM compliance_index ci
        JOIN countries c ON ci.country_code = c.code
        LEFT JOIN source_artifacts sa ON sa.id = ci.source_artifact_id
        WHERE 1=1
    """
    params = []
    if country_code: sql += " AND ci.country_code=%s"; params.append(country_code)
    if entry_type:   sql += " AND ci.entry_type=%s";   params.append(entry_type)
    if mandatory:    sql += " AND ci.mandatory=%s";    params.append(mandatory)
    if status:       sql += " AND ci.status=%s";       params.append(status)
    if product_code: sql += " AND %s=ANY(ci.applicable_products)"; params.append(product_code)
    if keyword:      sql += " AND ci.name ILIKE %s";   params.append(f"%{keyword}%")
    # Excel 是对外交付物，固定只导出已经官方证据闭环的记录。
    # 保留 verified 参数仅为兼容旧前端，不允许通过参数导出 candidate/suspicious。
    sql += " AND ci.authenticity_status = 'verified'"
    sql += " ORDER BY c.priority, ci.country_code, ci.entry_type, ci.mandatory DESC"

    with get_cursor() as cur:
        cur.execute(sql, params)
        rows = [dict(r) for r in cur.fetchall()]

    # 生成 Excel
    wb = openpyxl.Workbook()
    ws = getattr(wb, "active", None) or wb.create_sheet("合规知识库", 0)
    ws.title = "合规知识库"
    ws.freeze_panes = "A2"

    def hfill(c): return PatternFill("solid", fgColor=c)
    def border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    headers = [
        "认证/法规名称", "条目类型", "强制性", "国家", "优先级",
        "认证机构", "生效日期", "发布日期", "适用产品",
        "状态", "证据状态", "证据风险", "已核实", "官方链接",
        "原文工件", "文档ID", "摘要", "最后刷新"
    ]
    widths = [45, 10, 10, 12, 8, 24, 12, 12, 30, 8, 10, 10, 8, 40, 40, 36, 40, 18]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
        cell.fill = hfill("1F4E79")
        cell.border = border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    type_zh = {"regulation":"法规","standard":"标准","certification":"认证"}
    mand_zh = {"mandatory":"强制","voluntary":"自愿","recommended":"推荐"}

    for i, r in enumerate(rows, 2):
        # 未核实且置信度<80的行标黄
        bg = "FFF2CC" if not r.get("verified") and (r.get("confidence_score") or 100) < 80 else "FFFFFF"
        row_data = [
            r.get("name",""),
            type_zh.get(r.get("entry_type",""), r.get("entry_type","")),
            mand_zh.get(r.get("mandatory",""), r.get("mandatory","")),
            f"{r.get('country_name','')} ({r.get('country_code','')})",
            r.get("priority",""),
            r.get("issuing_body","") or "",
            str(r["effective_date"]) if r.get("effective_date") else "",
            str(r["published_date"]) if r.get("published_date") else "",
            "、".join(r.get("applicable_products") or []),
            r.get("status",""),
            r.get("authenticity_status",""),
            r.get("authenticity_risk_score",""),
            r.get("confidence_score",""),
            "✅" if r.get("verified") else "❌",
            r.get("official_url","") or "",
            r.get("source_artifact_url","") or "",
            str(r.get("document_id") or ""),
            r.get("summary","") or "",
            str(r["updated_at"])[:19] if r.get("updated_at") else "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(name="微软雅黑", size=9)
            cell.border = border()
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.fill = hfill(bg)
        ws.row_dimensions[i].height = 45

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    if hasattr(wb, "save"):
        wb.save(buf)
    else:
        buf.write(b"")
    buf.seek(0)

    today = date.today().strftime("%Y%m%d")
    filename = f"cybersec_compliance_{today}.xlsx"
    headers_resp = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    logger.info("导出Excel: %d条记录 by %s", len(rows), current_user)
    return StreamingResponse(buf, headers=headers_resp, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/review/pending")
async def get_pending_review_list(
    limit: int = 50,
    review_bucket: Optional[str] = Query(None),
    current_user: str = Depends(require_admin_user),
):
    """获取待人工核实的条目：优先看 candidate/suspicious，排除 quarantined/verified。"""
    bucket_clause = _pending_review_bucket_clause(review_bucket)
    with get_cursor() as cur:
        sql = """
            SELECT
              ci.compliance_id AS id,
              ci.name,
              ci.country_code,
              ci.entry_type,
              ci.mandatory,
              ci.status,
              ci.issuing_body,
              ci.official_url,
              ci.document_id AS source_document_id,
              ci.source_artifact_id,
              ci.authenticity_status,
              ci.authenticity_risk_score,
              GREATEST(0, 100 - COALESCE(ci.authenticity_risk_score, 0)) AS confidence_score,
              ci.applicable_products,
              ci.effective_date,
              ci.published_date,
              ci.updated_at,
              c.name_zh AS country_name,
              c.priority,
              rc.evidence_note AS authenticity_evidence,
              rc.reasons AS authenticity_reasons,
              rc.source_download_status,
              rc.source_download_error,
              COALESCE(sr.discovery_method, 'read_model') AS data_source
            FROM compliance_index ci
            JOIN countries c ON ci.country_code = c.code
            LEFT JOIN review_cases rc ON rc.compliance_id = ci.compliance_id
            LEFT JOIN source_records sr ON sr.id = ci.source_record_id
            WHERE ci.authenticity_status IN ('candidate', 'suspicious')
              AND ci.status = 'active'
        """
        if bucket_clause:
            sql += bucket_clause
        sql += """
            ORDER BY
              CASE COALESCE(ci.authenticity_status, 'candidate')
                WHEN 'suspicious' THEN 0
                ELSE 1
              END,
              CASE COALESCE(rc.source_download_status, 'pending')
                WHEN 'failed' THEN 0
                WHEN 'pending' THEN 1
                ELSE 2
              END,
              ci.authenticity_risk_score DESC NULLS LAST,
              c.priority
            LIMIT %s
        """
        cur.execute(sql, (limit,))
        rows = []
        for r in cur.fetchall():
            rows.append(_serialize_record_dates(dict(r)))
    return {"items": rows, "total": len(rows)}


@router.post("/", status_code=201)
async def create_compliance(
    data: ComplianceCreate,
    current_user: str = Depends(require_admin_user),
):
    """手动新增合规条目"""
    entry = data.model_dump(exclude_none=True)
    entry["data_source"] = f"manual:{current_user}"
    entry["verified"] = False
    entry["confidence_score"] = 60

    new_id = ComplianceRepository.create(entry)
    ComplianceRepository.set_authenticity_review(
        new_id,
        authenticity_status="candidate",
        risk_score=50,
        reasons=["manual_entry_requires_official_evidence"],
        checked_by=current_user,
        evidence="后台手动新增仅进入候选池；必须通过 manual-source 上传/确认官方正文页或 PDF 后才能标记 verified。",
    )
    record = ComplianceRepository.get_by_id(new_id)
    if record:
        ComplianceIndexRepository.refresh_for_compliance(record)
    ChangeLogRepository.record_change(
        record_id=new_id,
        change_type="created",
        old_value=None,
        new_value=entry,
        data_source=f"manual:{current_user}",
    )
    logger.info("手动新增: %s [by %s]", data.name, current_user)
    return {"id": new_id, "message": "创建成功"}


@router.get("/{record_id}")
async def get_compliance(
    record_id: str,
    current_user: str = Depends(get_current_user),
):
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT
              ci.compliance_id AS id,
              ci.name,
              ci.entry_type,
              ci.mandatory,
              ci.status,
              ci.country_code,
              c.name_zh AS country_name,
              c.priority,
              ci.issuing_body,
              ci.official_url,
              ci.applicable_products,
              ci.effective_date,
              ci.published_date,
              ci.summary,
              ci.summary AS scope_description,
              ci.summary AS remarks,
              ci.document_id AS source_document_id,
              ci.document_id,
              ci.source_artifact_id,
              ci.source_record_id,
              ci.canonical_requirement_id,
              ci.review_case_id,
              ci.authenticity_status,
              ci.authenticity_risk_score,
              GREATEST(0, 100 - COALESCE(ci.authenticity_risk_score, 0)) AS confidence_score,
              (ci.authenticity_status = 'verified') AS verified,
              ci.updated_at,
              COALESCE(sa.artifact_url, sa.official_url) AS source_artifact_url,
              sa.artifact_sha256 AS source_artifact_sha256,
              sa.download_status AS source_download_status,
              sa.download_error AS source_download_error,
              rc.evidence_note AS authenticity_evidence,
              rc.reasons AS authenticity_reasons
            FROM compliance_index ci
            JOIN countries c ON c.code = ci.country_code
            LEFT JOIN source_artifacts sa ON sa.id = ci.source_artifact_id
            LEFT JOIN review_cases rc ON rc.id = ci.review_case_id
            WHERE ci.compliance_id=%s
            LIMIT 1
            """,
            (record_id,),
        )
        index_row = cur.fetchone()
    if not index_row:
        raise HTTPException(status_code=404, detail="记录不存在")
    row = _serialize_record_dates(dict(index_row))
    if not is_admin_username(current_user) and row.get("authenticity_status") != "verified":
        raise HTTPException(status_code=404, detail="记录不存在")
    review_case = ReviewCaseRepository.get_by_compliance_id(record_id)
    canonical = CanonicalRequirementRepository.get_by_compliance_id(record_id)
    artifacts = SourceArtifactRepository.list_by_entity(record_id)
    evidence_payload = get_authenticity_review_service().get_evidence(record_id)

    row_translations: Dict[tuple[str, str], str] = {}
    row_translations.update(_translation_map("compliance_index", [record_id]))
    row = attach_translation_fields(row, row_translations, entity_id_field="id")
    review_payload = _serialize_record_dates(dict(review_case), ["checked_at", "created_at", "updated_at"]) if review_case else None
    if review_payload:
        review_payload = attach_translation_fields(
            review_payload,
            _translation_map("review_cases", [str(review_payload.get("id"))]),
            entity_id_field="id",
        )
    return {
        **row,
        "review_case": review_payload,
        "canonical_requirement": _serialize_record_dates(dict(canonical), ["created_at", "updated_at"]) if canonical else None,
        "source_artifacts": [_serialize_record_dates(dict(item), ["downloaded_at", "created_at", "updated_at"]) for item in artifacts],
        "evidence_events": [
            _serialize_record_dates(dict(item), ["created_at"])
            for item in (evidence_payload.get("events") or [])
        ],
    }


@router.post("/{record_id}/verify")
async def verify_record(
    record_id: str,
    current_user: str = Depends(require_admin_user),
):
    """旧入口已停用，避免在没有官方证据时盲目标 verified。"""
    raise HTTPException(
        status_code=409,
        detail="直接 /verify 已停用；verified 请使用 /manual-source，人工结论请使用 /review",
    )


@router.put("/{record_id}")
async def update_compliance(
    record_id: str,
    data: ComplianceUpdate,
    current_user: str = Depends(require_admin_user),
):
    """手动编辑合规条目"""
    old = ComplianceRepository.get_by_id(record_id)
    if not old:
        raise HTTPException(status_code=404, detail="记录不存在")

    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["data_source"] = f"manual:{current_user}"
    update_data.pop("verified", None)

    ok = ComplianceRepository.update(record_id, update_data, force=True)
    if ok:
        record = ComplianceRepository.get_by_id(record_id)
        if record:
            ComplianceIndexRepository.refresh_for_compliance(record)
        ChangeLogRepository.record_change(
            record_id=record_id,
            change_type="updated",
            old_value=dict(old),
            new_value=update_data,
            changed_fields=list(update_data.keys()),
            diff_summary=f"手动编辑 by {current_user}",
            data_source=f"manual:{current_user}",
        )
    logger.info("手动更新: %s [by %s]", record_id, current_user)
    return {"message": "更新成功"}


@router.delete("/{record_id}")
async def delete_compliance(
    record_id: str,
    current_user: str = Depends(require_admin_user),
):
    """废止合规条目（软删除，改状态为 deprecated）"""
    old = ComplianceRepository.get_by_id(record_id)
    if not old:
        raise HTTPException(status_code=404, detail="记录不存在")

    ComplianceRepository.deprecate(record_id)
    ChangeLogRepository.record_change(
        record_id=record_id,
        change_type="deprecated",
        old_value=dict(old),
        new_value={"status": "deprecated"},
        data_source=f"manual:{current_user}",
    )
    logger.info("手动废止: %s [by %s]", old["name"], current_user)
    return {"message": "已废止"}
