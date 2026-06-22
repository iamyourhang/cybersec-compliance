"""
admin/api/routes/models.py
产品型号管理接口 - 独立模块
"""
from __future__ import annotations
import io
import logging
from datetime import date
from typing import Any, Dict, List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from admin.api.auth import require_admin_user
from database.connection import get_cursor, get_connection

logger = logging.getLogger(__name__)
router = APIRouter()


# ---- Schema ----

class ModelCreate(BaseModel):
    code: str
    name: str
    name_en: Optional[str] = None
    category_code: str          # products.code，如 switch
    brand: Optional[str] = None
    description: Optional[str] = None
    specifications: Optional[Dict[str, Any]] = None

class ModelUpdate(BaseModel):
    name: Optional[str] = None
    name_en: Optional[str] = None
    brand: Optional[str] = None
    description: Optional[str] = None
    specifications: Optional[Dict[str, Any]] = None
    enabled: Optional[bool] = None

class OverrideCreate(BaseModel):
    compliance_id: str
    override_type: str = "include"   # include / exclude / customize
    custom_notes: Optional[str] = None


def _serialize_model(r: dict) -> dict:
    row = dict(r)
    for f in ["created_at", "updated_at"]:
        if row.get(f): row[f] = str(row[f])
    return row


# ---- 型号 CRUD ----

@router.get("/")
async def list_models(
    category_code: Optional[str] = None,
    brand: Optional[str] = None,
    keyword: Optional[str] = None,
    enabled: Optional[bool] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: str = Depends(require_admin_user),
):
    """列出产品型号"""
    sql = """
        SELECT pm.*, p.code AS category_code, p.name_zh AS category_name
        FROM product_models pm
        JOIN products p ON pm.category_id = p.id
        WHERE 1=1
    """
    params = []
    if category_code: sql += " AND p.code=%s"; params.append(category_code)
    if brand:         sql += " AND pm.brand ILIKE %s"; params.append(f"%{brand}%")
    if keyword:       sql += " AND (pm.code ILIKE %s OR pm.name ILIKE %s)"; params += [f"%{keyword}%"]*2
    if enabled is not None: sql += " AND pm.enabled=%s"; params.append(enabled)

    with get_cursor() as cur:
        cur.execute(f"SELECT COUNT(*) AS cnt FROM ({sql}) t", params)
        total = cur.fetchone()["cnt"]
        sql += " ORDER BY pm.brand, pm.code LIMIT %s OFFSET %s"
        params += [page_size, (page-1)*page_size]
        cur.execute(sql, params)
        items = [_serialize_model(r) for r in cur.fetchall()]

    return {"total": total, "page": page, "page_size": page_size, "items": items}


@router.post("/", status_code=201)
async def create_model(
    data: ModelCreate,
    current_user: str = Depends(require_admin_user),
):
    """新增产品型号"""
    # 查找大类 ID
    with get_cursor() as cur:
        cur.execute("SELECT id FROM products WHERE code=%s", (data.category_code,))
        row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=400, detail=f"产品大类不存在: {data.category_code}")
    category_id = row["id"]

    import json
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO product_models
                    (code, name, name_en, category_id, brand, description, specifications)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
            """, (
                data.code.upper(), data.name, data.name_en,
                category_id, data.brand, data.description,
                json.dumps(data.specifications) if data.specifications else None,
            ))
            new_id = str(cur.fetchone()[0])

    logger.info("➕ 新增型号: %s [%s] by %s", data.code, data.category_code, current_user)
    return {"id": new_id, "message": "创建成功"}


@router.get("/categories")
async def get_categories(current_user: str = Depends(require_admin_user)):
    """获取产品大类列表（供前端下拉）"""
    with get_cursor() as cur:
        cur.execute("""
            SELECT p.code, p.name_zh, p.name_en,
                   COUNT(pm.id) AS model_count
            FROM products p
            LEFT JOIN product_models pm ON pm.category_id = p.id AND pm.enabled=TRUE
            WHERE p.enabled=TRUE
            GROUP BY p.code, p.name_zh, p.name_en
            ORDER BY p.name_zh
        """)
        return [dict(r) for r in cur.fetchall()]


@router.get("/{model_id}")
async def get_model(
    model_id: str,
    current_user: str = Depends(require_admin_user),
):
    """获取型号详情"""
    with get_cursor() as cur:
        cur.execute("""
            SELECT pm.*, p.code AS category_code, p.name_zh AS category_name
            FROM product_models pm
            JOIN products p ON pm.category_id = p.id
            WHERE pm.id=%s
        """, (model_id,))
        row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="型号不存在")
    return _serialize_model(row)


@router.put("/{model_id}")
async def update_model(
    model_id: str,
    data: ModelUpdate,
    current_user: str = Depends(require_admin_user),
):
    """更新型号信息"""
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        return {"message": "无变更"}
    import json
    if "specifications" in updates:
        updates["specifications"] = json.dumps(updates["specifications"])
    set_clause = ", ".join(f"{k}=%s" for k in updates)
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE product_models SET {set_clause} WHERE id=%s",
                list(updates.values()) + [model_id]
            )
    return {"message": "更新成功"}


@router.delete("/{model_id}")
async def delete_model(
    model_id: str,
    current_user: str = Depends(require_admin_user),
):
    """删除型号（软删除）"""
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE product_models SET enabled=FALSE WHERE id=%s", (model_id,))
    return {"message": "已删除"}


# ---- 型号合规要求 ----

@router.get("/{model_id}/compliance")
async def get_model_compliance(
    model_id: str,
    country_code: Optional[str] = None,
    mandatory_only: bool = False,
    current_user: str = Depends(require_admin_user),
):
    """查询型号的完整合规要求（继承大类 + 定制化）"""
    sql = "SELECT * FROM v_model_compliance WHERE model_id=%s"
    params = [model_id]
    if country_code: sql += " AND country_code=%s"; params.append(country_code)
    if mandatory_only: sql += " AND mandatory='mandatory'"

    with get_cursor() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()

    items = []
    for r in rows:
        row = dict(r)
        if row.get("effective_date"): row["effective_date"] = str(row["effective_date"])
        items.append(row)

    return {
        "model_id": model_id,
        "total": len(items),
        "by_country": _group_by_country(items),
        "items": items,
    }


@router.post("/{model_id}/compliance/override")
async def add_compliance_override(
    model_id: str,
    data: OverrideCreate,
    current_user: str = Depends(require_admin_user),
):
    """为型号添加合规定制（排除/包含/自定义备注）"""
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO model_compliance_overrides
                    (model_id, compliance_id, override_type, custom_notes)
                VALUES (%s,%s,%s,%s)
                ON CONFLICT (model_id, compliance_id)
                DO UPDATE SET override_type=EXCLUDED.override_type,
                              custom_notes=EXCLUDED.custom_notes
            """, (model_id, data.compliance_id, data.override_type, data.custom_notes))
    return {"message": "定制化已保存"}


@router.delete("/{model_id}/compliance/override/{compliance_id}")
async def remove_compliance_override(
    model_id: str,
    compliance_id: str,
    current_user: str = Depends(require_admin_user),
):
    """删除型号合规定制，恢复继承大类"""
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM model_compliance_overrides WHERE model_id=%s AND compliance_id=%s",
                (model_id, compliance_id)
            )
    return {"message": "已恢复继承"}


# ---- 导出合规清单 ----

@router.get("/{model_id}/export")
async def export_model_compliance(
    model_id: str,
    country_code: Optional[str] = None,
    mandatory_only: bool = False,
    current_user: str = Depends(require_admin_user),
):
    """导出型号出口到指定国家的合规清单 Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 获取型号信息
    with get_cursor() as cur:
        cur.execute("""
            SELECT pm.*, p.name_zh AS category_name
            FROM product_models pm JOIN products p ON pm.category_id=p.id
            WHERE pm.id=%s
        """, (model_id,))
        model = cur.fetchone()
    if not model:
        raise HTTPException(status_code=404, detail="型号不存在")

    # 获取合规要求
    sql = "SELECT * FROM v_model_compliance WHERE model_id=%s"
    params = [model_id]
    if country_code: sql += " AND country_code=%s"; params.append(country_code)
    if mandatory_only: sql += " AND mandatory='mandatory'"
    sql += " ORDER BY priority, country_code, mandatory DESC"

    with get_cursor() as cur:
        cur.execute(sql, params)
        rows = [dict(r) for r in cur.fetchall()]

    # 生成 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合规清单"
    ws.freeze_panes = "A2"

    def hfill(c): return PatternFill("solid", fgColor=c)
    def bd():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    # 封面信息
    ws.cell(row=1, column=1, value=f"产品型号：{model['name']} ({model['code']})").font = Font(name="微软雅黑", bold=True, size=12, color="1F4E79")
    ws.cell(row=2, column=1, value=f"产品大类：{model['category_name']}  |  品牌：{model.get('brand','—')}  |  生成日期：{date.today().isoformat()}  |  合规条目：{len(rows)}条")
    ws.cell(row=2, column=1).font = Font(name="微软雅黑", size=10, color="666666")
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    # 表头
    headers = ["认证/法规名称", "类型", "强制性", "国家/地区", "优先级", "生效日期", "来源", "官方链接"]
    widths   = [50, 10, 10, 16, 8, 12, 12, 45]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
        cell.fill = hfill("1F4E79")
        cell.border = bd()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22

    type_zh = {"regulation":"法规","standard":"标准","certification":"认证"}
    mand_zh = {"mandatory":"强制","voluntary":"自愿","recommended":"推荐"}
    mand_color = {"mandatory":"FFE0E0","voluntary":"E2EFDA","recommended":"F0F0F0"}

    for i, r in enumerate(rows, 4):
        bg = mand_color.get(r.get("mandatory",""), "FFFFFF")
        row_data = [
            r.get("compliance_name",""),
            type_zh.get(r.get("entry_type",""), ""),
            mand_zh.get(r.get("mandatory",""), ""),
            f"{r.get('country_name','')} ({r.get('country_code','')})",
            r.get("priority",""),
            str(r["effective_date"]) if r.get("effective_date") else "—",
            "定制" if r.get("source") == "include" else "继承",
            r.get("official_url","") or "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(name="微软雅黑", size=9)
            cell.border = bd()
            cell.fill = hfill(bg)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 40

    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}3"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    country_str = f"_{country_code}" if country_code else "_全球"
    filename = f"{model['code']}_合规清单{country_str}_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def _group_by_country(items: List[dict]) -> Dict[str, Any]:
    result = {}
    for item in items:
        cc = item["country_code"]
        if cc not in result:
            result[cc] = {"country_name": item["country_name"], "priority": item["priority"], "items": []}
        result[cc]["items"].append(item)
    return result
