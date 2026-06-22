"""
collector/document/spec_excel_writer.py
规格要求 Excel 生成器 - 只负责把结构化数据写成 Excel
无 AI 依赖，无数据库依赖
"""
from __future__ import annotations
import io
import logging
from datetime import date
from typing import Any, Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# 颜色配置
COLORS = {
    "header_bg":   "1F4E79",
    "module_bg":   "2E75B6",
    "p1_bg":       "FFE0E0",
    "p2_bg":       "FFF2CC",
    "p3_bg":       "E2EFDA",
    "mandatory_bg":"FFE0E0",
    "recommended_bg":"E2EFDA",
    "white":       "FFFFFF",
    "light_gray":  "F5F5F5",
}

MODULE_COLORS = {
    "身份认证与密码管理": "D6E4F0",
    "网络访问控制":       "D6F0E4",
    "数据加密与传输安全": "F0E4D6",
    "安全更新与补丁管理": "E4D6F0",
    "日志与审计":         "F0F0D6",
    "漏洞管理与披露":     "F0D6D6",
    "物理与接口安全":     "D6F0F0",
    "安全配置与加固":     "F0D6F0",
    "供应链安全":         "D6D6F0",
    "合规认证与测试":     "F0E8D6",
}

def _hfont(bold=True, size=10, color="FFFFFF"):
    return Font(name="微软雅黑", bold=bold, size=size, color=color)

def _bfont(bold=False, size=9):
    return Font(name="微软雅黑", bold=bold, size=size, color="000000")

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _apply_header(ws, row, cols_values, widths):
    for col, (val, width) in enumerate(zip(cols_values, widths), 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _hfont()
        cell.fill = _fill(COLORS["header_bg"])
        cell.border = _border()
        cell.alignment = _center()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 24


def generate_spec_excel(
    specs: List[Dict[str, Any]],
    regulation_name: str,
    country_code: str,
    output_path: str = None,
) -> bytes:
    """
    生成规格要求 Excel。
    返回 Excel 字节内容，同时可选写入文件。
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet1: 封面信息
    _build_cover_sheet(wb, regulation_name, country_code, specs)
    # Sheet2: 全量规格列表（双语）
    _build_full_sheet(wb, specs, regulation_name, country_code)
    # Sheet3: 按模块分组
    _build_module_sheets(wb, specs)
    # Sheet4: P1强制要求（销售必须满足）
    _build_priority_sheet(wb, specs, "P1", "P1-销售必须满足", COLORS["p1_bg"])

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    if output_path:
        with open(output_path, "wb") as f:
            f.write(content)
        logger.info("✅ Excel已保存: %s (%dKB, %d条规格)", output_path, len(content)//1024, len(specs))

    return content


def _build_cover_sheet(wb, regulation_name, country_code, specs):
    ws = wb.create_sheet("📋 文档信息")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60

    info = [
        ("文档标题", f"{regulation_name} - 产品规格要求"),
        ("Title", f"{regulation_name} - Product Security Requirements"),
        ("适用法规", regulation_name),
        ("国家/地区", country_code),
        ("生成日期", date.today().isoformat()),
        ("规格总数", len(specs)),
        ("P1强制规格", len([s for s in specs if s.get("priority") == "P1"])),
        ("P2重要规格", len([s for s in specs if s.get("priority") == "P2"])),
        ("P3推荐规格", len([s for s in specs if s.get("priority") == "P3"])),
        ("功能模块数", len(set(s.get("module_zh","") for s in specs))),
    ]

    ws.cell(row=1, column=1, value="网安合规助手 · 产品规格要求文档").font = Font(name="微软雅黑", bold=True, size=16, color="1F4E79")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 40

    for i, (k, v) in enumerate(info, 3):
        kc = ws.cell(row=i, column=1, value=k)
        vc = ws.cell(row=i, column=2, value=v)
        kc.font = _bfont(bold=True)
        kc.fill = _fill("EAF2FB")
        kc.border = _border()
        kc.alignment = _left()
        vc.font = _bfont()
        vc.border = _border()
        vc.alignment = _left()
        ws.row_dimensions[i].height = 20

    # 模块统计
    from collections import Counter
    module_counts = Counter(s.get("module_zh", "未分类") for s in specs)
    ws.cell(row=len(info)+4, column=1, value="模块分布").font = _bfont(bold=True, size=10)
    for j, (mod, cnt) in enumerate(sorted(module_counts.items()), len(info)+5):
        ws.cell(row=j, column=1, value=mod).font = _bfont()
        ws.cell(row=j, column=2, value=f"{cnt} 条").font = _bfont()
        ws.row_dimensions[j].height = 18


def _build_full_sheet(wb, specs, regulation_name, country_code):
    ws = wb.create_sheet("📄 全量规格列表")
    ws.freeze_panes = "A2"

    headers = [
        "规格ID", "功能模块(中)", "功能模块(英)", "规格标题(中)", "规格标题(英)",
        "规格描述(中)", "规格描述(英)", "适用产品", "强制性",
        "法规条款", "验证方法(中)", "验证方法(英)", "优先级", "备注(中)"
    ]
    widths = [12, 16, 20, 22, 25, 45, 45, 30, 10, 15, 35, 35, 8, 30]
    _apply_header(ws, 1, headers, widths)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    mandatory_label = {"mandatory": "强制", "recommended": "推荐"}
    for i, spec in enumerate(specs, 2):
        module = spec.get("module_zh", "")
        bg = MODULE_COLORS.get(module, COLORS["light_gray"])
        priority = spec.get("priority", "P2")
        p_color = {"P1": COLORS["p1_bg"], "P2": COLORS["p2_bg"], "P3": COLORS["p3_bg"]}.get(priority, COLORS["white"])

        row_data = [
            spec.get("req_id", ""),
            module,
            spec.get("module_en", ""),
            spec.get("title_zh", ""),
            spec.get("title_en", ""),
            spec.get("description_zh", ""),
            spec.get("description_en", ""),
            "、".join(spec.get("applicable_products", [])),
            mandatory_label.get(spec.get("mandatory", ""), spec.get("mandatory", "")),
            spec.get("regulation_clause", ""),
            spec.get("verification_method_zh", ""),
            spec.get("verification_method_en", ""),
            priority,
            spec.get("notes_zh", ""),
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _bfont()
            cell.border = _border()
            cell.alignment = _left()
            # 优先级列用优先级颜色，其他用模块颜色
            if col == 13:
                cell.fill = _fill(p_color)
                cell.alignment = _center()
                cell.font = _bfont(bold=True)
            elif col in (1, 9):
                cell.fill = _fill(bg)
                cell.alignment = _center()
            else:
                cell.fill = _fill(COLORS["white"])
        ws.row_dimensions[i].height = 55


def _build_module_sheets(wb, specs):
    """按功能模块各建一个Sheet"""
    from collections import defaultdict
    by_module = defaultdict(list)
    for s in specs:
        by_module[s.get("module_zh", "未分类")].append(s)

    for module_zh, module_specs in sorted(by_module.items()):
        sheet_name = module_zh[:31]
        ws = wb.create_sheet(sheet_name)
        ws.freeze_panes = "A2"

        headers = ["规格ID", "规格标题(中/英)", "规格描述(中)", "规格描述(英)", "适用产品", "验证方法(中)", "优先级"]
        widths  = [12, 28, 50, 50, 28, 40, 8]
        _apply_header(ws, 1, headers, widths)

        bg = MODULE_COLORS.get(module_zh, COLORS["light_gray"])
        for i, spec in enumerate(module_specs, 2):
            priority = spec.get("priority", "P2")
            p_color = {"P1": COLORS["p1_bg"], "P2": COLORS["p2_bg"], "P3": COLORS["p3_bg"]}.get(priority, COLORS["white"])
            title_bilingual = f"{spec.get('title_zh','')}\n{spec.get('title_en','')}"
            row_data = [
                spec.get("req_id", ""),
                title_bilingual,
                spec.get("description_zh", ""),
                spec.get("description_en", ""),
                "、".join(spec.get("applicable_products", [])),
                spec.get("verification_method_zh", ""),
                priority,
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = _bfont()
                cell.border = _border()
                cell.alignment = _left()
                if col in (1, 7):
                    cell.fill = _fill(bg if col == 1 else p_color)
                    cell.alignment = _center()
                    if col == 7: cell.font = _bfont(bold=True)
                else:
                    cell.fill = _fill(COLORS["white"])
            ws.row_dimensions[i].height = 60


def _build_priority_sheet(wb, specs, priority, sheet_name, bg_color):
    """按优先级筛选建Sheet"""
    filtered = [s for s in specs if s.get("priority") == priority]
    if not filtered:
        return
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A2"
    headers = ["规格ID", "功能模块", "规格标题(中)", "规格描述(中)", "规格描述(英)", "法规条款", "验证方法(中)"]
    widths  = [12, 16, 22, 50, 50, 15, 40]
    _apply_header(ws, 1, headers, widths)
    for i, spec in enumerate(filtered, 2):
        row_data = [
            spec.get("req_id", ""),
            spec.get("module_zh", ""),
            spec.get("title_zh", ""),
            spec.get("description_zh", ""),
            spec.get("description_en", ""),
            spec.get("regulation_clause", ""),
            spec.get("verification_method_zh", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _bfont(bold=(col in (1,3)))
            cell.border = _border()
            cell.fill = _fill(bg_color if col == 1 else COLORS["white"])
            cell.alignment = _left()
        ws.row_dimensions[i].height = 60
