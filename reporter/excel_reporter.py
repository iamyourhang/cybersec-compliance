"""
reporter/excel_reporter.py
Excel 周报生成器 - 总览矩阵 + 认证明细 + 按产品分 Sheet
"""

from __future__ import annotations

import io
import logging
from datetime import date, datetime, timezone
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from database.connection import get_cursor
from database.repository import ComplianceLifecycleRepository

logger = logging.getLogger(__name__)

# ---- 样式常量 ----
COLOR_HEADER = "1F4E79"      # 深蓝（表头）
COLOR_SUBHEADER = "2E75B6"   # 中蓝（子表头）
COLOR_MANDATORY = "FFF2CC"   # 浅黄（强制）
COLOR_VOLUNTARY = "E2EFDA"   # 浅绿（自愿）
COLOR_DANGER = "FFE0E0"      # 浅红（即将生效/已过期）
COLOR_REGULATION = "DDEEFF"  # 浅蓝（法规行）
COLOR_CERT = "F0F0F0"        # 浅灰（认证行）
WHITE = "FFFFFF"
BLACK = "000000"


def _header_font(bold=True) -> Font:
    return Font(name="微软雅黑", bold=bold, color=WHITE, size=10)

def _body_font(bold=False) -> Font:
    return Font(name="微软雅黑", bold=bold, color=BLACK, size=9)

def _header_fill(color=COLOR_HEADER) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _border() -> Border:
    side = Side(style="thin", color="BBBBBB")
    return Border(left=side, right=side, top=side, bottom=side)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _apply_header_row(ws, row_idx: int, values: List[str], widths: List[int]) -> None:
    """应用表头行样式"""
    for col, (val, width) in enumerate(zip(values, widths), 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.border = _border()
        cell.alignment = _center()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row_idx].height = 22


class ExcelReporter:
    """生成合规知识库 Excel 周报"""

    def generate(self, output_path: Optional[str] = None) -> bytes:
        """
        生成 Excel 报告。
        output_path 指定时同时写入文件。
        返回 Excel 字节内容。
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认空 Sheet

        logger.info("开始生成 Excel 报告...")

        # 1. 总览矩阵 Sheet
        self._build_overview_sheet(wb)
        # 2. 认证明细 Sheet
        self._build_detail_sheet(wb)
        # 3. 即将生效预警 Sheet
        self._build_upcoming_sheet(wb)
        # 4. 本周变更 Sheet
        self._build_changes_sheet(wb)
        # 5. 按产品分 Sheet
        self._build_product_sheets(wb)

        # 保存
        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()

        if output_path:
            with open(output_path, "wb") as f:
                f.write(content)
            logger.info("✅ Excel 报告已保存: %s (%d KB)", output_path, len(content) // 1024)

        return content

    # --------------------------------------------------------
    # Sheet 1：总览矩阵（国家 × 产品）
    # --------------------------------------------------------

    def _build_overview_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("总览矩阵", 0)
        ws.freeze_panes = "B2"

        # 获取数据
        with get_cursor() as cur:
            cur.execute("""
                SELECT c.name_zh AS country, c.priority,
                       p.name_zh AS product, p.code AS product_code,
                       COUNT(ci.id) FILTER (WHERE ci.mandatory='mandatory') AS mandatory_cnt,
                       COUNT(ci.id) FILTER (WHERE ci.mandatory='voluntary') AS voluntary_cnt,
                       COUNT(ci.id) AS total_cnt
                FROM countries c
                CROSS JOIN products p
                LEFT JOIN compliance_index ci
                    ON ci.country_code = c.code
                    AND p.code = ANY(ci.applicable_products)
                    AND ci.status = 'active'
                    AND ci.authenticity_status = 'verified'
                WHERE c.enabled = TRUE AND p.enabled = TRUE
                GROUP BY c.name_zh, c.priority, p.name_zh, p.code
                ORDER BY c.priority, c.name_zh, p.name_zh
            """)
            rows = cur.fetchall()

        # 获取产品列表（列）
        with get_cursor() as cur:
            cur.execute("SELECT code, name_zh FROM products WHERE enabled=TRUE ORDER BY name_zh")
            products = cur.fetchall()

        product_codes = [p["code"] for p in products]
        product_names = [p["name_zh"] for p in products]

        # 构建矩阵数据
        matrix: Dict[str, Dict[str, str]] = {}
        for row in rows:
            country = row["country"]
            if country not in matrix:
                matrix[country] = {"_priority": row["priority"]}
            pcode = row["product_code"]
            if row["total_cnt"] > 0:
                matrix[country][pcode] = f"强制{row['mandatory_cnt']}/自愿{row['voluntary_cnt']}"
            else:
                matrix[country][pcode] = "—"

        # 写表头
        header = ["国家/地区", "优先级"] + product_names
        widths = [16, 8] + [14] * len(product_names)
        _apply_header_row(ws, 1, header, widths)

        # 写数据行
        row_idx = 2
        for country, data in sorted(matrix.items(), key=lambda x: (x[1].get("_priority", "P3"), x[0])):
            ws.cell(row=row_idx, column=1, value=country).font = _body_font(bold=True)
            ws.cell(row=row_idx, column=2, value=data.get("_priority", "")).alignment = _center()
            for col, pcode in enumerate(product_codes, 3):
                val = data.get(pcode, "—")
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = _body_font()
                cell.alignment = _center()
                cell.border = _border()
                if "强制" in str(val) and val != "—":
                    cell.fill = PatternFill("solid", fgColor=COLOR_MANDATORY)
            ws.cell(row=row_idx, column=1).border = _border()
            ws.cell(row=row_idx, column=2).border = _border()
            row_idx += 1

        ws.title = "📊 总览矩阵"
        logger.debug("总览矩阵 Sheet 完成，%d 个国家", len(matrix))

    # --------------------------------------------------------
    # Sheet 2：认证明细
    # --------------------------------------------------------

    def _build_detail_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("认证明细")
        ws.freeze_panes = "A2"

        headers = [
            "认证/法规名称", "条目类型", "强制/自愿", "国家/地区", "认证机构",
            "生效日期", "发布日期", "适用产品", "状态", "证据状态",
            "证据风险", "官方链接", "原文工件", "文档ID", "摘要", "最后刷新"
        ]
        widths = [40, 10, 10, 10, 20, 12, 12, 30, 8, 10, 10, 35, 40, 36, 40, 18]
        _apply_header_row(ws, 1, headers, widths)

        with get_cursor() as cur:
            cur.execute("""
                SELECT ci.name, ci.entry_type, ci.mandatory, ci.country_code,
                       c.name_zh AS country_name, ci.issuing_body,
                       ci.effective_date, ci.published_date,
                       ci.applicable_products, ci.status, ci.authenticity_status,
                       ci.authenticity_risk_score, ci.official_url,
                       COALESCE(sa.artifact_url, sa.official_url) AS artifact_url,
                       ci.document_id, ci.summary, ci.updated_at
                FROM compliance_index ci
                JOIN countries c ON ci.country_code = c.code
                LEFT JOIN source_artifacts sa ON sa.id = ci.source_artifact_id
                WHERE ci.status = 'active'
                  AND ci.authenticity_status='verified'
                ORDER BY c.priority, ci.country_code, ci.entry_type, ci.mandatory DESC
            """)
            records = cur.fetchall()

        entry_type_zh = {"regulation": "法规", "standard": "标准", "certification": "认证"}
        mandatory_zh = {"mandatory": "强制", "voluntary": "自愿", "recommended": "推荐"}

        row_idx = 2
        for rec in records:
            fill_color = COLOR_REGULATION if rec["entry_type"] == "regulation" else COLOR_CERT
            row_data = [
                rec["name"],
                entry_type_zh.get(rec["entry_type"], rec["entry_type"]),
                mandatory_zh.get(rec["mandatory"], rec["mandatory"]),
                f"{rec['country_name']} ({rec['country_code']})",
                rec["issuing_body"] or "",
                str(rec["effective_date"]) if rec["effective_date"] else "",
                str(rec["published_date"]) if rec["published_date"] else "",
                "、".join(rec["applicable_products"] or []),
                rec["status"],
                rec["authenticity_status"],
                rec["authenticity_risk_score"],
                rec["official_url"] or "",
                rec["artifact_url"] or "",
                str(rec["document_id"]) if rec["document_id"] else "",
                rec["summary"] or "",
                str(rec["updated_at"])[:19] if rec["updated_at"] else "",
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = _body_font()
                cell.border = _border()
                cell.alignment = _left()
                cell.fill = PatternFill("solid", fgColor=fill_color)
                if col == 6 and rec.get("effective_date"):
                    # 生效日期临近着色
                    days_left = (rec["effective_date"] - date.today()).days
                    if days_left <= 30:
                        cell.fill = PatternFill("solid", fgColor=COLOR_DANGER)

            row_idx += 1

        ws.title = "📋 认证明细"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        logger.debug("认证明细 Sheet 完成，%d 条记录", len(records))

    # --------------------------------------------------------
    # Sheet 3：即将生效预警
    # --------------------------------------------------------

    def _build_upcoming_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("⚠️ 即将生效")
        ws.freeze_panes = "A2"

        headers = ["认证/法规名称", "国家/地区", "优先级", "节点", "节点日期", "距今(天)", "强制/自愿", "适用产品", "官方链接"]
        widths = [40, 14, 8, 24, 12, 10, 10, 30, 40]
        _apply_header_row(ws, 1, headers, widths)

        records = ComplianceLifecycleRepository.get_upcoming_milestones(days=30, limit=500)

        row_idx = 2
        for rec in records:
            days = rec["days_until_effective"]
            fill_color = COLOR_DANGER if days <= 7 else COLOR_MANDATORY

            row_data = [
                rec["name"],
                f"{rec['country_name']} ({rec['country_code']})",
                rec["priority"],
                rec.get("milestone_label_zh") or "生效/适用节点",
                str(rec["effective_date"]),
                days,
                rec["mandatory"],
                "、".join(rec.get("applicable_products") or []),
                rec.get("official_url") or "",
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = _body_font(bold=(days <= 7))
                cell.border = _border()
                cell.alignment = _center() if col in (3, 6, 7) else _left()
                cell.fill = PatternFill("solid", fgColor=fill_color)
            row_idx += 1

        logger.debug("即将生效 Sheet: %d 条预警", len(records))

    # --------------------------------------------------------
    # Sheet 4：本周变更
    # --------------------------------------------------------

    def _build_changes_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("本周变更")
        ws.freeze_panes = "A2"

        headers = ["变更时间", "变更类型", "认证/法规名称", "国家", "变更字段", "变更摘要", "是否已审核"]
        widths = [18, 10, 40, 12, 20, 50, 10]
        _apply_header_row(ws, 1, headers, widths)

        with get_cursor() as cur:
            cur.execute("""
                WITH events AS (
                    SELECT cl.changed_at,
                           cl.change_type::TEXT AS change_type,
                           cl.changed_fields,
                           cl.diff_summary,
                           cl.reviewed,
                           ci.name,
                           ci.country_code,
                           c.name_zh AS country_name,
                           0 AS source_rank
                    FROM change_log cl
                    JOIN compliance_index ci ON cl.record_id = ci.compliance_id
                    JOIN countries c ON ci.country_code = c.code
                    WHERE cl.changed_at >= NOW() - INTERVAL '7 days'
                      AND ci.status='active'
                      AND ci.authenticity_status='verified'
                    UNION ALL
                    SELECT rc.checked_at AS changed_at,
                           'created' AS change_type,
                           ARRAY['authenticity_status']::TEXT[] AS changed_fields,
                           '官方证据核验通过，进入正式知识库' AS diff_summary,
                           TRUE AS reviewed,
                           ci.name,
                           ci.country_code,
                           c.name_zh AS country_name,
                           1 AS source_rank
                    FROM compliance_index ci
                    JOIN review_cases rc ON rc.id = ci.review_case_id
                    JOIN countries c ON ci.country_code = c.code
                    WHERE rc.checked_at >= NOW() - INTERVAL '7 days'
                      AND rc.current_status='verified'
                      AND ci.status='active'
                      AND ci.authenticity_status='verified'
                )
                SELECT changed_at, change_type, changed_fields, diff_summary,
                       reviewed, name, country_code, country_name
                FROM (
                    SELECT DISTINCT ON (name, country_code)
                           changed_at, change_type, changed_fields, diff_summary,
                           reviewed, name, country_code, country_name, source_rank
                    FROM events
                    WHERE changed_at IS NOT NULL
                    ORDER BY name, country_code, changed_at DESC, source_rank
                ) latest
                ORDER BY changed_at DESC
                LIMIT 200
            """)
            changes = cur.fetchall()

        change_type_zh = {"created": "新增", "updated": "更新", "deprecated": "废止"}
        row_idx = 2
        for ch in changes:
            row_data = [
                str(ch["changed_at"])[:19] if ch["changed_at"] else "",
                change_type_zh.get(ch["change_type"], ch["change_type"]),
                ch["name"],
                f"{ch['country_name']} ({ch['country_code']})",
                "、".join(ch["changed_fields"] or []),
                ch["diff_summary"] or "",
                "✅" if ch["reviewed"] else "待审核",
            ]
            fill = COLOR_VOLUNTARY if ch["reviewed"] else COLOR_MANDATORY
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = _body_font()
                cell.border = _border()
                cell.alignment = _left()
                cell.fill = PatternFill("solid", fgColor=fill)
            row_idx += 1

        logger.debug("本周变更 Sheet: %d 条", len(changes))

    # --------------------------------------------------------
    # Sheet 5+：按产品分 Sheet
    # --------------------------------------------------------

    def _build_product_sheets(self, wb: openpyxl.Workbook) -> None:
        with get_cursor() as cur:
            cur.execute("SELECT code, name_zh FROM products WHERE enabled=TRUE ORDER BY name_zh")
            products = cur.fetchall()

        for product in products:
            self._build_single_product_sheet(wb, product["code"], product["name_zh"])

    def _build_single_product_sheet(self, wb: openpyxl.Workbook, product_code: str, product_name: str) -> None:
        ws = wb.create_sheet(product_name.replace('/', '_').replace('\\', '_')[:31])
        ws.freeze_panes = "A2"

        headers = ["认证/法规名称", "条目类型", "强制/自愿", "国家/地区", "生效日期", "认证机构", "证据状态", "官方链接", "摘要"]
        widths = [40, 10, 10, 14, 12, 24, 10, 40, 40]
        _apply_header_row(ws, 1, headers, widths)

        with get_cursor() as cur:
            cur.execute("""
                SELECT ci.name, ci.entry_type, ci.mandatory, ci.country_code,
                       c.name_zh AS country_name, c.priority,
                       ci.effective_date, ci.issuing_body,
                       ci.authenticity_status, ci.official_url, ci.summary
                FROM compliance_index ci
                JOIN countries c ON ci.country_code = c.code
                WHERE %s = ANY(ci.applicable_products)
                  AND ci.status = 'active'
                  AND ci.authenticity_status='verified'
                ORDER BY c.priority, ci.mandatory DESC, ci.name
            """, (product_code,))
            records = cur.fetchall()

        entry_type_zh = {"regulation": "法规", "standard": "标准", "certification": "认证"}
        mandatory_zh = {"mandatory": "强制", "voluntary": "自愿", "recommended": "推荐"}

        row_idx = 2
        for rec in records:
            row_data = [
                rec["name"],
                entry_type_zh.get(rec["entry_type"], ""),
                mandatory_zh.get(rec["mandatory"], ""),
                f"{rec['country_name']} ({rec['country_code']})",
                str(rec["effective_date"]) if rec["effective_date"] else "",
                rec["issuing_body"] or "",
                rec["authenticity_status"],
                rec["official_url"] or "",
                rec["summary"] or "",
            ]
            fill_color = COLOR_MANDATORY if rec["mandatory"] == "mandatory" else COLOR_VOLUNTARY
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = _body_font()
                cell.border = _border()
                cell.alignment = _left()
                cell.fill = PatternFill("solid", fgColor=fill_color)
            row_idx += 1

        logger.debug("产品 Sheet [%s]: %d 条", product_name, len(records))
