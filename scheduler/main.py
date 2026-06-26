"""
scheduler/main.py
APScheduler 调度器主程序
- 每天凌晨：关键法规倒计时刷新、官方源/原文/解析/规格/预警分层任务
- 每周：AI 官方候选发现
- 每两周：证据驱动完整闭环生成 Excel 并发送飞书
"""

from __future__ import annotations

import logging
import signal
import sys
import json
import os
import tempfile
from math import ceil
from pathlib import Path
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.logger import setup_logging
from config.settings import get_settings
from database.connection import get_cursor, get_connection, health_check
from database.repository import ComplianceLifecycleRepository

setup_logging(level="INFO")
logger = logging.getLogger(__name__)
BIWEEKLY_UPDATE_START = datetime(2026, 5, 4, 1, 0, tzinfo=ZoneInfo("Asia/Shanghai"))
MONTHLY_AI_DISCOVERY_PRIORITIES = ["P1", "P2", "P3"]
MONTHLY_AI_DISCOVERY_LIMIT_COUNTRIES = 320
MONTHLY_AI_DISCOVERY_QUERIES_PER_COUNTRY = 6
MONTHLY_AI_DISCOVERY_CRON = "30 0 1 * *"
MONTHLY_AI_DISCOVERY_REPORT_LOOKBACK_DAYS = 31
MONTHLY_AI_DISCOVERY_REPORT_LIMIT = 5000
WEEKLY_FRONTLINE_DIGEST_CRON = "0 9 * * 1"
WEEKLY_FRONTLINE_DIGEST_LOOKBACK_HOURS = 24 * 7
WEEKLY_FRONTLINE_DIGEST_LIMIT = 30

# Backward-compatible aliases for manual callers and older tests/scripts.
WEEKLY_AI_DISCOVERY_PRIORITIES = MONTHLY_AI_DISCOVERY_PRIORITIES
WEEKLY_AI_DISCOVERY_LIMIT_COUNTRIES = MONTHLY_AI_DISCOVERY_LIMIT_COUNTRIES
WEEKLY_AI_DISCOVERY_QUERIES_PER_COUNTRY = MONTHLY_AI_DISCOVERY_QUERIES_PER_COUNTRY
DAILY_AI_DISCOVERY_PRIORITIES = MONTHLY_AI_DISCOVERY_PRIORITIES
DAILY_AI_DISCOVERY_LIMIT_COUNTRIES = MONTHLY_AI_DISCOVERY_LIMIT_COUNTRIES
DAILY_AI_DISCOVERY_QUERIES_PER_COUNTRY = MONTHLY_AI_DISCOVERY_QUERIES_PER_COUNTRY


def get_official_source_pipeline():
    from collector.official_sources.pipeline import get_official_source_pipeline as _factory

    return _factory()


# ============================================================
# 任务函数
# ============================================================

def job_official_source_sync_daily() -> None:
    """每日同步全球官方源，作为日报的确定性主来源。"""
    logger.info("⏰ [调度] 开始每日官方源同步（P1/P2/P3）")
    stats = get_official_source_pipeline().sync_country_priorities(["P1", "P2", "P3"])
    logger.info("⏰ [调度] 每日官方源同步完成: %s", stats)


def job_official_source_sync_weekly() -> None:
    """每周同步 P2/P3 国家官方源"""
    logger.info("⏰ [调度] 开始每周官方源同步（P2/P3）")
    stats = get_official_source_pipeline().sync_country_priorities(["P2", "P3"])
    logger.info("⏰ [调度] 每周官方源同步完成: %s", stats)


def job_official_artifact_fetch(limit: int = 20) -> dict:
    """抓取官方原文工件到 COS"""
    logger.info("⏰ [调度] 开始抓取官方原文工件")
    from collector.document.source_ingest import OfficialSourceIngestService
    from database.repository import SourceArtifactRepository, SourceRecordRepository

    service = OfficialSourceIngestService()
    items = SourceRecordRepository.list_pending_artifact_records(limit=limit)
    success = 0
    for item in items:
        try:
            service.ingest_record(item, requested_by="system:official_source")
            success += 1
        except Exception as exc:
            logger.warning("工件抓取失败 [%s]: %s", item.get("name"), exc)
            SourceArtifactRepository.upsert_for_compliance(
                compliance_id=item.get("compliance_id"),
                official_url=item.get("source_url"),
                artifact_url=item.get("artifact_url") or item.get("source_url"),
                artifact_type=None,
                artifact_sha256=None,
                download_status="failed",
                download_error=str(exc),
                source_record_id=str(item["id"]),
            )
    result = {"success": success, "total": len(items)}
    logger.info("⏰ [调度] 官方原文工件抓取完成: %s", result)
    return result


def job_candidate_verification(limit: int = 50) -> dict:
    """为已有关联条目的 source candidates 建立审核桶，不自动放行 verified"""
    logger.info("⏰ [调度] 开始候选审核分桶")
    from database.repository import (
        CanonicalRequirementRepository,
        ComplianceIndexRepository,
        ComplianceRepository,
        ReviewCaseRepository,
        SourceRecordRepository,
    )

    items = SourceRecordRepository.list_bucketable_records(limit=limit)
    bucketed = 0
    for item in items:
        compliance_id = item.get("compliance_id")
        if not compliance_id:
            continue
        record = ComplianceRepository.get_by_id(str(compliance_id))
        if not record:
            continue
        ReviewCaseRepository.ensure_for_record(record)
        CanonicalRequirementRepository.upsert_from_compliance(
            record,
            verification_status="verified"
            if (record.get("authenticity_status") or "candidate") == "verified"
            else "candidate",
        )
        ComplianceIndexRepository.refresh_for_compliance(record)
        bucketed += 1
    result = {"bucketed": bucketed, "total": len(items)}
    logger.info("⏰ [调度] 候选审核分桶完成: %s", result)
    return result


def job_document_parse(limit: int = 10) -> dict:
    """解析已抓取的官方原文文档"""
    logger.info("⏰ [调度] 开始解析官方原文文档")
    from collector.document.doc_repository import DocRepository
    from admin.api.routes.documents import _parse_and_index_document

    docs = DocRepository.list_pending_source_documents(limit=limit)
    for doc in docs:
        _parse_and_index_document(str(doc["id"]), False)
    result = {"total": len(docs)}
    logger.info("⏰ [调度] 官方原文文档解析完成: %s", result)
    return result


def _list_verified_documents_needing_specs(limit: int = 10) -> list[dict]:
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT d.id
            FROM regulation_documents d
            JOIN compliance_index ci ON ci.compliance_id = d.compliance_id
            WHERE d.index_status='ready'
              AND d.parse_status='done'
              AND ci.status='active'
              AND ci.authenticity_status='verified'
              AND COALESCE(d.spec_requirement_count, 0) = 0
              AND d.cos_key IS NOT NULL
            ORDER BY d.indexed_at ASC NULLS LAST, d.created_at ASC
            LIMIT %s
            """,
            (limit,),
        )
        return [dict(row) for row in cur.fetchall()]


def job_spec_generate(limit: int = 10) -> dict:
    """只对 verified 官方原文生成规格要求。"""
    logger.info("⏰ [调度] 开始生成 verified 原文规格")
    from collector.document.spec_generator import SpecGeneratorService

    docs = _list_verified_documents_needing_specs(limit)
    service = SpecGeneratorService()
    generated = 0
    failed = 0
    for doc in docs:
        try:
            service.generate_from_doc(str(doc["id"]))
            generated += 1
        except Exception as exc:
            failed += 1
            logger.warning("规格生成失败 [%s]: %s", doc.get("id"), exc)
    result = {"total": len(docs), "generated": generated, "failed": failed}
    logger.info("⏰ [调度] verified 原文规格生成完成: %s", result)
    return result


def job_read_model_refresh(limit: int = 200) -> dict:
    """刷新兼容读模型，供后台/飞书/RAG 统一消费"""
    logger.info("⏰ [调度] 开始刷新读模型")
    from database.connection import get_cursor
    from database.repository import ComplianceIndexRepository, ComplianceRepository

    with get_cursor() as cur:
        cur.execute(
            """
            SELECT id
            FROM compliance_knowledge
            ORDER BY updated_at DESC
            LIMIT %s
            """,
            (limit,),
        )
        ids = [str(row["id"]) for row in cur.fetchall()]

    refreshed = 0
    for compliance_id in ids:
        record = ComplianceRepository.get_by_id(compliance_id)
        if not record:
            continue
        ComplianceIndexRepository.refresh_for_compliance(record)
        refreshed += 1
    result = {"refreshed": refreshed, "total": len(ids)}
    logger.info("⏰ [调度] 读模型刷新完成: %s", result)
    return result


def job_global_source_registry_refresh() -> dict:
    """刷新官方源种子和逐国覆盖矩阵，不直接创建 verified 记录。"""
    logger.info("⏰ [调度] 开始刷新全球官方源注册表")
    from scripts.refresh_country_source_coverage import refresh as refresh_coverage
    from scripts.seed_country_research_outcomes import seed as seed_research_outcomes
    from scripts.seed_full_country_catalog import seed as seed_country_catalog
    from scripts.seed_global_official_sources import seed as seed_global_sources

    country_catalog = seed_country_catalog()
    inserted, updated = seed_global_sources()
    research_outcomes = seed_research_outcomes()
    coverage = refresh_coverage()
    result = {
        "country_catalog": country_catalog,
        "official_sources_inserted": inserted,
        "official_sources_updated": updated,
        "research_outcomes_updated": research_outcomes,
        "coverage": coverage,
    }
    logger.info("⏰ [调度] 全球官方源注册表刷新完成: %s", result)
    return result


def job_key_regulation_countdown_refresh() -> dict:
    """每日刷新关键法规分阶段适用节点，供倒计时、看板和飞书预警使用。"""
    logger.info("⏰ [调度] 开始刷新关键法规适用节点倒计时")
    try:
        result = ComplianceLifecycleRepository.seed_key_regulation_milestones()
        logger.info("⏰ [调度] 关键法规适用节点倒计时刷新完成: %s", result)
        return result
    except Exception as exc:
        logger.error("⏰ [调度] 关键法规适用节点倒计时刷新失败: %s", exc, exc_info=True)
        return {"status": "failed", "error": str(exc)}


def job_monthly_ai_discovery(limit_countries: int = MONTHLY_AI_DISCOVERY_LIMIT_COUNTRIES) -> dict:
    """每月受控 AI 发现：只生成和校验官方候选 source_records，不直接 verified。"""
    logger.info("⏰ [调度] 开始每月 AI 官方候选发现")
    try:
        from collector.discovery.service import get_ai_discovery_service

        result = get_ai_discovery_service().run(
            priorities=MONTHLY_AI_DISCOVERY_PRIORITIES,
            limit_countries=limit_countries,
            queries_per_country=MONTHLY_AI_DISCOVERY_QUERIES_PER_COUNTRY,
            validation_mode="ai",
        )
        result.update(_send_monthly_ai_discovery_report(result))
        logger.info("⏰ [调度] 每月 AI 官方候选发现完成: %s", result)
        return result
    except Exception as exc:
        logger.error("⏰ [调度] 每月 AI 官方候选发现失败: %s", exc, exc_info=True)
        return {
            "status": "failed",
            "candidate_count": 0,
            "accepted_count": 0,
            "rejected_count": 0,
            "error": str(exc),
        }


def _send_monthly_ai_discovery_report(result: dict) -> dict:
    """导出月度 AI 发现总表，并向飞书发送下载入口。"""
    report_result = {
        "report_sent": False,
        "report_cos_url": None,
        "report_file_name": None,
        "report_row_count": 0,
    }
    tmp_path: Path | None = None
    try:
        from notifier.feishu import get_notifier

        rows = _collect_monthly_ai_discovery_report_rows(
            run_id=result.get("run_id"),
            limit=MONTHLY_AI_DISCOVERY_REPORT_LIMIT,
        )
        report_result["report_row_count"] = len(rows)
        generated_at = datetime.now(ZoneInfo("Asia/Shanghai"))
        filename = f"ai_discovery_{generated_at.strftime('%Y%m%d_%H%M%S')}.xlsx"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)
        _write_ai_discovery_report_excel(rows, tmp_path)

        settings = get_settings()
        cos_url = None
        if settings.cos.secret_id:
            cos_url = _upload_to_cos(str(tmp_path), filename, settings)
        report_result["report_cos_url"] = cos_url
        report_result["report_file_name"] = filename

        notifier = get_notifier()
        if notifier:
            report_result["report_sent"] = notifier.send_ai_discovery_report_card(
                candidate_count=int(result.get("candidate_count") or 0),
                accepted_count=int(result.get("accepted_count") or 0),
                rejected_count=int(result.get("rejected_count") or 0),
                reference_count=int(result.get("reference_count") or 0),
                report_row_count=len(rows),
                report_url=cos_url,
                generated_at=generated_at.strftime("%Y-%m-%d %H:%M"),
            )
        return report_result
    except Exception as exc:
        logger.warning("每月 AI 发现报告生成/推送失败: %s", exc, exc_info=True)
        report_result["report_error"] = str(exc)
        return report_result
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except FileNotFoundError:
                pass


def _collect_monthly_ai_discovery_report_rows(run_id: str | None = None, limit: int = MONTHLY_AI_DISCOVERY_REPORT_LIMIT) -> list[dict]:
    """收集最近一次月度窗口内的 AI 官方发现记录。

    source_records 暂未持久化 run_id，因此这里以最近 31 天被 AI discovery 创建或更新的记录作为月度总表。
    """
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT sr.id::TEXT AS id,
                   sr.country_code,
                   COALESCE(c.name_zh, sr.country_code) AS country_name,
                   sr.title,
                   sr.entry_type,
                   sr.source_status,
                   sr.source_url,
                   sr.artifact_url,
                   sr.published_date,
                   sr.created_at,
                   sr.updated_at,
                   sr.source_payload,
                   sa.download_status,
                   sa.download_error
            FROM source_records sr
            LEFT JOIN countries c ON c.code = sr.country_code
            LEFT JOIN source_artifacts sa ON sa.source_record_id = sr.id
            WHERE sr.discovery_method = 'ai_weekly_discovery'
              AND sr.updated_at >= NOW() - make_interval(days => %s)
            ORDER BY sr.updated_at DESC, sr.created_at DESC
            LIMIT %s
            """,
            (MONTHLY_AI_DISCOVERY_REPORT_LOOKBACK_DAYS, limit),
        )
        return [dict(row) for row in cur.fetchall()]


def _write_ai_discovery_report_excel(rows: list[dict], output_path: Path) -> Path:
    """生成 AI 发现候选总表。"""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI发现候选"
    headers = [
        "国家/地区",
        "国家代码",
        "类型",
        "状态",
        "标题",
        "中文标题",
        "摘要/说明",
        "官方原文链接",
        "工件链接",
        "发布日期",
        "发现/更新时间",
        "AI相关性理由",
        "官方性理由",
        "下载状态",
        "下载错误",
    ]
    ws.append(headers)

    for row in rows:
        payload = _coerce_json_dict(row.get("source_payload"))
        raw_candidate = _coerce_json_dict(payload.get("raw_candidate"))
        ws.append([
            row.get("country_name") or row.get("country_code") or "",
            row.get("country_code") or "",
            _entry_type_label(row.get("entry_type")),
            _source_status_label(row.get("source_status")),
            row.get("title") or "",
            raw_candidate.get("title_zh") or raw_candidate.get("name_zh") or "",
            raw_candidate.get("summary_zh") or raw_candidate.get("summary") or payload.get("ai_reason") or "",
            row.get("source_url") or "",
            row.get("artifact_url") or "",
            _format_cell_date(row.get("published_date")),
            _format_cell_date(row.get("updated_at") or row.get("created_at")),
            payload.get("cyber_product_relevance_reason") or "",
            payload.get("official_evidence_reason") or "",
            row.get("download_status") or "",
            row.get("download_error") or "",
        ])

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [16, 10, 14, 14, 34, 34, 52, 52, 52, 14, 18, 46, 46, 14, 36]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(output_path)
    return output_path


def _coerce_json_dict(value) -> dict:
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip():
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return parsed if isinstance(parsed, dict) else {}
    return {}


def _format_cell_date(value) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.astimezone(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d %H:%M")
    return str(value)


def _entry_type_label(value) -> str:
    return {
        "regulation": "法律法规",
        "certification": "认证",
        "standard": "标准",
        "scheme": "认证方案",
        "guideline": "指南",
    }.get(str(value or ""), str(value or ""))


def _source_status_label(value) -> str:
    return {
        "candidate": "待审核候选",
        "validation_pending": "待AI/人工校验",
        "reference": "动态参考",
        "rejected": "已拒绝",
    }.get(str(value or ""), str(value or ""))


def job_weekly_ai_discovery(limit_countries: int = MONTHLY_AI_DISCOVERY_LIMIT_COUNTRIES) -> dict:
    """Compatibility wrapper; scheduled AI discovery now runs monthly."""
    return job_monthly_ai_discovery(limit_countries=limit_countries)


def job_daily_ai_discovery(limit_countries: int = MONTHLY_AI_DISCOVERY_LIMIT_COUNTRIES) -> dict:
    """Compatibility wrapper; scheduled AI discovery now runs monthly."""
    return job_monthly_ai_discovery(limit_countries=limit_countries)


def job_weekly_compliance_update() -> dict:
    """每两周官方证据驱动更新闭环：源注册、发现、抓原文、分桶、解析、刷新、发报告。"""
    logger.info("⏰ [调度] 开始每两周全球合规知识库更新闭环")
    from collector.workflow.evidence_pipeline import EvidencePipelineRunner

    registry_result = job_global_source_registry_refresh()
    ai_discovery_result = {
        "status": "scheduled_separately",
        "cadence_days": 30,
        "job_id": "monthly_ai_discovery",
    }
    runner = EvidencePipelineRunner(
        source_sync=lambda priorities: get_official_source_pipeline().sync_country_priorities(list(priorities)),
        artifact_fetch=job_official_artifact_fetch,
        review_bucket=job_candidate_verification,
        document_parse=job_document_parse,
        spec_generate=job_spec_generate,
        read_model_refresh=job_read_model_refresh,
        weekly_report=job_weekly_report,
    )
    result = runner.run_weekly_closed_loop()
    result["source_collection"]["source_registry_refresh"] = registry_result
    result["source_collection"]["ai_discovery"] = ai_discovery_result
    logger.info("⏰ [调度] 每两周全球合规知识库更新闭环完成: %s", result)
    return result


def job_alert_scan() -> None:
    """预警扫描任务"""
    logger.info("⏰ [调度] 开始预警扫描")
    try:
        from notifier.alert_scanner import AlertScanner
        scanner = AlertScanner()
        result = scanner.run()
        logger.info("⏰ [调度] 预警扫描完成: %s", result)
    except Exception as e:
        logger.error("⏰ [调度] 预警扫描失败: %s", e, exc_info=True)
        raise


def job_frontline_feishu_digest() -> dict:
    """手动生成网安合规摘要；默认不再每日自动推送。"""
    logger.info("⏰ [手动] 开始网安合规摘要")
    try:
        from notifier.alert_scanner import AlertScanner
        scanner = AlertScanner()
        sent = scanner.scan_frontline_digest(lookback_hours=_hours_since_local_midnight())
        result = {"sent": bool(sent), "count": sent}
        logger.info("⏰ [手动] 网安合规摘要完成: %s", result)
        return result
    except Exception as e:
        logger.error("⏰ [手动] 网安合规摘要失败: %s", e, exc_info=True)
        raise


def job_weekly_frontline_feishu_digest() -> dict:
    """每周网安合规动态：汇总最近7天线索、入库和适用节点后推送飞书。"""
    logger.info("⏰ [调度] 开始每周网安合规动态")
    try:
        from notifier.alert_scanner import AlertScanner
        scanner = AlertScanner()
        sent = scanner.scan_frontline_digest(
            lookback_hours=WEEKLY_FRONTLINE_DIGEST_LOOKBACK_HOURS,
            limit=WEEKLY_FRONTLINE_DIGEST_LIMIT,
        )
        result = {"sent": bool(sent), "count": sent}
        logger.info("⏰ [调度] 每周网安合规动态完成: %s", result)
        return result
    except Exception as e:
        logger.error("⏰ [调度] 每周网安合规动态失败: %s", e, exc_info=True)
        raise


def _hours_since_local_midnight(now: datetime | None = None) -> int:
    now = now or datetime.now(ZoneInfo("Asia/Shanghai"))
    local_now = now.astimezone(ZoneInfo("Asia/Shanghai"))
    midnight = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
    return max(1, int(ceil((local_now - midnight).total_seconds() / 3600)))


def job_weekly_report() -> dict:
    """周报生成 + 飞书发送任务"""
    logger.info("⏰ [调度] 开始生成周报")
    try:
        from reporter.excel_reporter import ExcelReporter
        from notifier.feishu import get_notifier
        import tempfile, os

        settings = get_settings()

        # 生成 Excel
        reporter = ExcelReporter()
        report_date = datetime.now(timezone.utc).strftime("%Y%m%d")
        filename = f"cybersec_compliance_{report_date}.xlsx"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        content = reporter.generate(output_path=tmp_path)

        # 上传 COS（如已配置）
        cos_url = None
        if settings.cos.secret_id:
            try:
                cos_url = _upload_to_cos(tmp_path, filename, settings)
            except Exception as e:
                logger.warning("COS 上传失败: %s", e)

        os.unlink(tmp_path)

        stats = _collect_weekly_report_stats()

        # 发飞书
        notifier = get_notifier()
        feishu_sent = False
        if notifier:
            feishu_sent = notifier.send_weekly_report_card(
                total_records=stats["total_records"],
                country_count=stats["country_count"],
                candidate_this_week=stats["candidate_this_week"],
                verified_this_week=stats["verified_this_week"],
                source_artifacts_this_week=stats["source_artifacts_this_week"],
                quarantined_this_week=stats["quarantined_this_week"],
                upcoming_alerts=stats["upcoming_alerts"],
                report_url=cos_url,
            )

        # 记录到 report_records 表
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO report_records
                        (report_type, report_date, file_name, cos_url, feishu_sent, feishu_sent_at, stats)
                    VALUES ('weekly', CURRENT_DATE, %s, %s, %s, NOW(), %s)
                    """,
                    (filename, cos_url, feishu_sent, json.dumps(stats, ensure_ascii=False, default=str)),
                )

        logger.info("⏰ [调度] 周报生成完成 [COS=%s]", cos_url or "未上传")
        return {"sent": feishu_sent, "cos_url": cos_url, "file_name": filename}

    except Exception as e:
        logger.error("⏰ [调度] 周报生成失败: %s", e, exc_info=True)
        raise


def _collect_weekly_report_stats() -> dict:
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT COUNT(*) AS total
            FROM compliance_index
            WHERE status='active' AND authenticity_status='verified'
            """
        )
        total = cur.fetchone()["total"]

        cur.execute(
            """
            SELECT COUNT(DISTINCT country_code) AS cnt
            FROM compliance_index
            WHERE status='active' AND authenticity_status='verified'
            """
        )
        country_cnt = cur.fetchone()["cnt"]

        cur.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM source_records
            WHERE source_status='candidate'
              AND created_at >= NOW() - INTERVAL '7 days'
            """
        )
        candidate_cnt = cur.fetchone()["cnt"]

        cur.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM review_cases
            WHERE current_status='verified'
              AND checked_at >= NOW() - INTERVAL '7 days'
            """
        )
        verified_cnt = cur.fetchone()["cnt"]

        cur.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM review_cases
            WHERE current_status='quarantined'
              AND checked_at >= NOW() - INTERVAL '7 days'
            """
        )
        quarantined_cnt = cur.fetchone()["cnt"]

        cur.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM source_artifacts
            WHERE download_status='downloaded'
              AND downloaded_at >= NOW() - INTERVAL '7 days'
            """
        )
        artifact_cnt = cur.fetchone()["cnt"]

        upcoming = ComplianceLifecycleRepository.get_upcoming_milestones(days=30, limit=10)

    return {
        "total_records": total,
        "country_count": country_cnt,
        "candidate_this_week": candidate_cnt,
        "verified_this_week": verified_cnt,
        "quarantined_this_week": quarantined_cnt,
        "source_artifacts_this_week": artifact_cnt,
        "upcoming_alerts": upcoming,
    }


def _upload_to_cos(local_path: str, filename: str, settings) -> str:
    """上传文件到腾讯云 COS，返回下载 URL"""
    from qcloud_cos import CosConfig, CosS3Client

    config = CosConfig(
        Region=settings.cos.region,
        SecretId=settings.cos.secret_id,
        SecretKey=settings.cos.secret_key,
    )
    client = CosS3Client(config)
    key = f"{settings.cos.report_prefix}{filename}"
    with open(local_path, "rb") as f:
        client.put_object(Bucket=settings.cos.bucket, Body=f, Key=key)
    return f"https://{settings.cos.bucket}.cos.{settings.cos.region}.myqcloud.com/{key}"


# ============================================================
# 调度器配置
# ============================================================

def build_scheduler() -> BlockingScheduler:
    from apscheduler.schedulers.blocking import BlockingScheduler
    from apscheduler.triggers.cron import CronTrigger
    from apscheduler.triggers.interval import IntervalTrigger

    settings = get_settings()
    scheduler = BlockingScheduler(timezone="Asia/Shanghai")

    # 全球官方源：每天凌晨1点；作为日报的确定性主来源
    scheduler.add_job(
        job_official_source_sync_daily,
        CronTrigger.from_crontab("0 1 * * *", timezone="Asia/Shanghai"),
        id="official_source_sync_daily",
        name="全球官方源同步",
        max_instances=1,
        misfire_grace_time=3600,
    )

    # 每两周完整更新闭环：官方源、原文、审核分桶、解析、读模型、Excel与飞书
    scheduler.add_job(
        job_weekly_compliance_update,
        IntervalTrigger(weeks=2, start_date=BIWEEKLY_UPDATE_START, timezone="Asia/Shanghai"),
        id="weekly_compliance_update",
        name="每两周全球合规知识库更新",
        max_instances=1,
        misfire_grace_time=7200,
    )

    # 关键法规倒计时：每天刷新 CRA 等分阶段适用节点，确保看板/飞书窗口准确
    scheduler.add_job(
        job_key_regulation_countdown_refresh,
        CronTrigger.from_crontab("5 0 * * *", timezone="Asia/Shanghai"),
        id="key_regulation_countdown_refresh",
        name="关键法规适用节点倒计时刷新",
        max_instances=1,
        misfire_grace_time=3600,
    )

    # AI 发现：每月1日凌晨0点30，只产生官方候选线索和待审核证据，不直接写正式库
    scheduler.add_job(
        job_monthly_ai_discovery,
        CronTrigger.from_crontab(MONTHLY_AI_DISCOVERY_CRON, timezone="Asia/Shanghai"),
        id="monthly_ai_discovery",
        name="每月AI官方候选发现",
        max_instances=1,
        misfire_grace_time=7200,
    )

    # 原文工件抓取：每天凌晨2点30
    scheduler.add_job(
        job_official_artifact_fetch,
        CronTrigger.from_crontab("30 2 * * *", timezone="Asia/Shanghai"),
        id="official_artifact_fetch",
        name="官方原文抓取",
        max_instances=1,
        misfire_grace_time=3600,
    )

    # 候选验真：每天凌晨3点
    scheduler.add_job(
        job_candidate_verification,
        CronTrigger.from_crontab("0 3 * * *", timezone="Asia/Shanghai"),
        id="candidate_verification",
        name="候选规则验真",
        max_instances=1,
        misfire_grace_time=3600,
    )

    # 文档解析：每天凌晨3点30分
    scheduler.add_job(
        job_document_parse,
        CronTrigger.from_crontab("30 3 * * *", timezone="Asia/Shanghai"),
        id="document_parse",
        name="官方原文解析",
        max_instances=1,
        misfire_grace_time=3600,
    )

    # 规格输出：只处理 verified 且已完成索引的官方原文
    scheduler.add_job(
        job_spec_generate,
        CronTrigger.from_crontab("45 3 * * *", timezone="Asia/Shanghai"),
        id="spec_generate",
        name="verified原文规格输出",
        max_instances=1,
        misfire_grace_time=3600,
    )

    # 预警扫描：每天凌晨4点
    scheduler.add_job(
        job_alert_scan,
        CronTrigger.from_crontab("0 4 * * *", timezone="Asia/Shanghai"),
        id="alert_scan",
        name="预警扫描",
        max_instances=1,
        misfire_grace_time=3600,
    )

    # 每周合规动态：每周一上午9点，汇总最近7天新增线索、正式入库和生效节点
    scheduler.add_job(
        job_weekly_frontline_feishu_digest,
        CronTrigger.from_crontab(WEEKLY_FRONTLINE_DIGEST_CRON, timezone="Asia/Shanghai"),
        id="weekly_frontline_feishu_digest",
        name="每周网安合规动态",
        max_instances=1,
        misfire_grace_time=3600,
    )

    # 覆盖矩阵刷新：每天凌晨4点30，仅统计，不改变真实性状态
    scheduler.add_job(
        job_global_source_registry_refresh,
        CronTrigger.from_crontab("30 4 * * *", timezone="Asia/Shanghai"),
        id="source_registry_refresh",
        name="全球官方源覆盖矩阵刷新",
        max_instances=1,
        misfire_grace_time=3600,
    )

    return scheduler


def _on_job_error(event) -> None:
    logger.error("❌ 任务 [%s] 执行失败: %s", event.job_id, event.exception)


def _on_job_executed(event) -> None:
    retval = getattr(event, "retval", None)
    if isinstance(retval, (int, float)):
        logger.info("✅ 任务 [%s] 执行完成，返回值: %.1f", event.job_id, float(retval))
    elif retval is None:
        logger.info("✅ 任务 [%s] 执行完成", event.job_id)
    else:
        logger.info("✅ 任务 [%s] 执行完成，返回: %s", event.job_id, retval)


def main() -> None:
    from apscheduler.events import EVENT_JOB_ERROR, EVENT_JOB_EXECUTED

    logger.info("=" * 60)
    logger.info("  网安合规助手 - 调度器启动")
    logger.info("=" * 60)

    db_health = health_check()
    if db_health["status"] != "healthy":
        logger.critical("❌ 数据库不可用，调度器退出")
        sys.exit(1)
    logger.info("✅ 数据库健康")

    scheduler = build_scheduler()
    scheduler.add_listener(_on_job_error, EVENT_JOB_ERROR)
    scheduler.add_listener(_on_job_executed, EVENT_JOB_EXECUTED)

    # 打印任务列表
    logger.info("已配置任务：")
    for job in scheduler.get_jobs():
        logger.info("  • [%s] %s", job.id, job.trigger)

    # 优雅退出
    def _shutdown(signum, frame):
        logger.info("收到信号 %s，正在优雅退出...", signum)
        scheduler.shutdown(wait=True)
        sys.exit(0)

    signal.signal(signal.SIGTERM, _shutdown)
    signal.signal(signal.SIGINT, _shutdown)

    logger.info("调度器运行中（Ctrl+C 退出）")
    scheduler.start()


if __name__ == "__main__":
    main()
