import sys
import types


class _Dim:
    width = None
    height = None


class _DimDict(dict):
    def __missing__(self, key):
        self[key] = _Dim()
        return self[key]


class _Cell:
    def __init__(self):
        self.value = None
        self.font = None
        self.fill = None
        self.border = None
        self.alignment = None


class _Worksheet:
    def __init__(self):
        self.freeze_panes = None
        self.column_dimensions = _DimDict()
        self.row_dimensions = _DimDict()
        self.title = ""
        self.auto_filter = type("_Filter", (), {"ref": ""})()
        self.cells = {}

    def cell(self, row, column, value=None):
        key = (row, column)
        cell = self.cells.setdefault(key, _Cell())
        if value is not None:
            cell.value = value
        return cell


class _Workbook:
    def create_sheet(self, name, index=None):
        return _Worksheet()


def _install_fake_openpyxl():
    openpyxl_mod = types.ModuleType("openpyxl")
    openpyxl_mod.Workbook = _Workbook
    styles_mod = types.ModuleType("openpyxl.styles")
    styles_mod.Alignment = lambda **kwargs: ("alignment", kwargs)
    styles_mod.Border = lambda **kwargs: ("border", kwargs)
    styles_mod.Font = lambda **kwargs: ("font", kwargs)
    styles_mod.PatternFill = lambda *args, **kwargs: ("fill", args, kwargs)
    styles_mod.Side = lambda **kwargs: ("side", kwargs)
    utils_mod = types.ModuleType("openpyxl.utils")
    utils_mod.get_column_letter = lambda col: str(col)
    sys.modules.setdefault("openpyxl", openpyxl_mod)
    sys.modules.setdefault("openpyxl.styles", styles_mod)
    sys.modules.setdefault("openpyxl.utils", utils_mod)


_install_fake_openpyxl()

import openpyxl
from reporter.excel_reporter import ExcelReporter


class _FakeCursor:
    def __init__(self):
        self.sqls = []
        self._last_sql = ""

    def execute(self, sql, params=None):
        self.sqls.append(sql)
        self._last_sql = sql

    def fetchall(self):
        return []


class _CursorContext:
    def __init__(self, cursor):
        self.cursor = cursor

    def __enter__(self):
        return self.cursor

    def __exit__(self, exc_type, exc, tb):
        return False


def test_detail_sheet_reads_verified_compliance_index(monkeypatch):
    cursor = _FakeCursor()
    monkeypatch.setattr("reporter.excel_reporter.get_cursor", lambda: _CursorContext(cursor))

    wb = openpyxl.Workbook()
    ExcelReporter()._build_detail_sheet(wb)

    detail_sql = "\n".join(cursor.sqls)
    assert "FROM compliance_index ci" in detail_sql
    assert "ci.authenticity_status='verified'" in detail_sql
    assert "JOIN compliance_knowledge" not in detail_sql


def test_overview_sheet_counts_verified_index_records(monkeypatch):
    cursor = _FakeCursor()
    monkeypatch.setattr("reporter.excel_reporter.get_cursor", lambda: _CursorContext(cursor))

    wb = openpyxl.Workbook()
    ExcelReporter()._build_overview_sheet(wb)

    overview_sql = "\n".join(cursor.sqls)
    assert "LEFT JOIN compliance_index ci" in overview_sql
    assert "ci.authenticity_status = 'verified'" in overview_sql


def test_upcoming_sheet_reads_verified_index_records(monkeypatch):
    cursor = _FakeCursor()
    monkeypatch.setattr("reporter.excel_reporter.get_cursor", lambda: _CursorContext(cursor))
    monkeypatch.setattr(
        "reporter.excel_reporter.ComplianceLifecycleRepository.get_upcoming_milestones",
        lambda days, limit: [
            {
                "name": "Cyber Resilience Act",
                "country_name": "欧盟",
                "country_code": "EU",
                "priority": "P0",
                "milestone_label_zh": "漏洞与严重事件报告义务开始适用",
                "effective_date": "2026-09-11",
                "days_until_effective": 119,
                "mandatory": "mandatory",
                "applicable_products": ["router"],
                "official_url": "https://eur-lex.europa.eu/",
            }
        ],
    )

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("placeholder")
    created = []
    original_create_sheet = wb.create_sheet

    def _capture_sheet(name, index=None):
        sheet = original_create_sheet(name, index)
        created.append(sheet)
        return sheet

    wb.create_sheet = _capture_sheet
    ExcelReporter()._build_upcoming_sheet(wb)

    upcoming_sheet = created[-1]
    assert upcoming_sheet.cells[(1, 4)].value == "节点"
    assert upcoming_sheet.cells[(2, 1)].value == "Cyber Resilience Act"
    assert upcoming_sheet.cells[(2, 4)].value == "漏洞与严重事件报告义务开始适用"
    assert cursor.sqls == []


def test_changes_sheet_includes_review_verified_events(monkeypatch):
    cursor = _FakeCursor()
    monkeypatch.setattr("reporter.excel_reporter.get_cursor", lambda: _CursorContext(cursor))

    wb = openpyxl.Workbook()
    ExcelReporter()._build_changes_sheet(wb)

    changes_sql = "\n".join(cursor.sqls)
    assert "FROM compliance_index ci" in changes_sql
    assert "review_cases rc" in changes_sql
    assert "ci.status='active'" in changes_sql
    assert "ci.authenticity_status='verified'" in changes_sql


def test_product_sheet_reads_verified_index_records(monkeypatch):
    cursor = _FakeCursor()
    monkeypatch.setattr("reporter.excel_reporter.get_cursor", lambda: _CursorContext(cursor))

    wb = openpyxl.Workbook()
    ExcelReporter()._build_single_product_sheet(wb, "router", "路由器")

    product_sql = "\n".join(cursor.sqls)
    assert "FROM compliance_index ci" in product_sql
    assert "ci.authenticity_status='verified'" in product_sql
    assert "ci.status = 'active'" in product_sql
    assert "JOIN compliance_knowledge" not in product_sql
